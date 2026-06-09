"""
openpyxl workbook builder — assembles all sheets including the mandatory audit log.

Amendment V1.2, Section 6.1 — Sheet 8 Mandatory Regulatory Audit Log:
  Every workbook produced by this builder MUST include the audit log sheet
  at position 8 (index 7).  The build() method enforces this unconditionally.

Amendment V1.2, Section 6.2 — Data Quality Watermarking:
  AI-extracted cells in any sheet receive a lineage comment:
      [Lineage Context: AI-Assisted Non-XBRL Extraction |
       Document Source: Page {page_number} |
       Confidence: {confidence_pct}%]
  Applied via services.export.sheets.sources.watermark_ai_cells().

Amendment V1.2, Section 5 — Export Blocking:
  The ValidationEngine.assert_exportable() guard MUST be called before
  any sheet is written.  If CRITICAL validation failures exist, build()
  raises ValidationBlockedError and no file is produced.

Sheet layout (10 sheets total):
  Sheet 1  — Overview
  Sheet 2  — Income Statement
  Sheet 3  — Balance Sheet
  Sheet 4  — Cash Flow
  Sheet 5  — Financial Ratios
  Sheet 6  — Segments
  Sheet 7  — Strategic Initiatives
  Sheet 8  — Audit Log (MANDATORY — Amendment V1.2 §6.1)
  Sheet 9  — Validation Report
  Sheet 10 — Metadata

Milestone: M6-Step62 (interface pre-provisioned by Amendment V1.2 compliance sweep)
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import TYPE_CHECKING

from services.export.sheets.sources import AuditRow, build_audit_log_sheet, watermark_ai_cells
from services.validation.engine import (
    FinancialDataBag,
    ValidationBlockedError,
    ValidationEngine,
    ValidationResult,
)

if TYPE_CHECKING:
    pass


# ---------------------------------------------------------------------------
# Workbook build request
# ---------------------------------------------------------------------------


@dataclass
class WorkbookBuildRequest:
    """
    Input to the workbook builder.

    Attributes:
        financial_bag:   Pre-signed, USD-normalised financial data for validation.
        audit_rows:      One AuditRow per extracted data point (for Sheet 8).
        company_name:    Display name for the workbook cover/header.
        fiscal_year:     Fiscal year being exported.
        fiscal_period:   Q1 | Q2 | Q3 | Q4 | FY.
        reporting_standard: US_GAAP | IFRS | IND_AS.
        ai_cell_map:     Mapping of sheet_name → {cell_coordinate → (page, confidence)}
                         for Amendment V1.2 §6.2 AI cell watermarking.
    """

    financial_bag: FinancialDataBag
    audit_rows: list[AuditRow] = field(default_factory=list)
    company_name: str = ""
    fiscal_year: int = 0
    fiscal_period: str = "FY"
    reporting_standard: str = "US_GAAP"
    ai_cell_map: dict[str, dict[str, tuple[int, float]]] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Builder
# ---------------------------------------------------------------------------


class WorkbookBuilder:
    """
    Assembles a compliant Excel workbook from extracted financial data.

    Enforces Amendment V1.2 §5 (export blocking on validation failure),
    §6.1 (Sheet 8 audit log), and §6.2 (AI cell watermarking).

    Usage::

        builder = WorkbookBuilder()
        wb_bytes = builder.build(request)
        with open("output.xlsx", "wb") as f:
            f.write(wb_bytes)
    """

    def __init__(self) -> None:
        self._validation_engine = ValidationEngine()

    def validate(self, request: WorkbookBuildRequest) -> ValidationResult:
        """
        Run the dual-dimension validation engine against the financial data.

        Returns the ValidationResult without raising even on CRITICAL findings.
        Call this if you want to inspect findings before deciding to export.

        Args:
            request: WorkbookBuildRequest containing the financial data bag.

        Returns:
            ValidationResult with all findings.
        """
        return self._validation_engine.run(request.financial_bag)

    def build(self, request: WorkbookBuildRequest) -> bytes:
        """
        Build and return the Excel workbook as bytes.

        Amendment V1.2 §5: Runs validation first. Raises ValidationBlockedError
        if CRITICAL failures exist — no file is produced in that case.

        Amendment V1.2 §6.1: Always writes the Sheet 8 audit log at position 8.

        Amendment V1.2 §6.2: Applies lineage comments to AI-extracted cells.

        Args:
            request: WorkbookBuildRequest.

        Returns:
            Raw bytes of the .xlsx file.

        Raises:
            ValidationBlockedError: When CRITICAL validation failures exist.
        """
        import io
        import openpyxl  # type: ignore[import]

        # Amendment V1.2 §5: validation gate — MUST run before any sheet write.
        result = self._validation_engine.run(request.financial_bag)
        self._validation_engine.assert_exportable(result)  # raises if CRITICAL

        wb = openpyxl.Workbook()

        # ── Sheets 1–7 (TODO M6-Step62) ──────────────────────────────────────
        # Each sheet builder will be implemented in M6. Placeholder sheets are
        # created here so position 8 is always the audit log.
        _PLACEHOLDER_NAMES = [
            "Overview",
            "Income Statement",
            "Balance Sheet",
            "Cash Flow",
            "Ratios",
            "Segments",
            "Strategic Initiatives",
        ]
        # Remove default sheet created by openpyxl.
        default_sheet = wb.active
        for name in _PLACEHOLDER_NAMES:
            wb.create_sheet(title=name)
        if default_sheet in wb.worksheets:
            del wb[default_sheet.title]

        # ── Sheet 8 — Mandatory Audit Log (Amendment V1.2 §6.1) ──────────────
        build_audit_log_sheet(wb, request.audit_rows)

        # ── Sheets 9–10 (TODO M6) ─────────────────────────────────────────────
        wb.create_sheet(title="Validation Report")
        wb.create_sheet(title="Metadata")

        # ── Amendment V1.2 §6.2 — AI cell watermarking ───────────────────────
        for sheet_name, cell_map in request.ai_cell_map.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for cell_coord, (page_num, confidence) in cell_map.items():
                    watermark_ai_cells(
                        ws,
                        cell_coord,
                        page_number=page_num,
                        confidence_pct=confidence,
                    )

        # ── Serialise to bytes ────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
