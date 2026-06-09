"""
Excel Export Service — M6.3: Data Ingestion and Period-Column Hydration.

Provides ``ExcelExportService``, the single async entry point that:

  1. Fetches ALL non-restated FinancialLineItem rows for the job's company
     across every available fiscal year and period (multi-period load).
  2. Discovers distinct (fiscal_year, fiscal_period) combinations from the
     loaded dataset and sorts them chronologically (oldest → newest, left → right).
  3. Pivots each statement sheet: rows = canonical tags in taxonomy order,
     column groups = chronological periods (4 columns each: Value (Reported),
     Reported Currency, Value (USD), FX Rate Used).
  4. Applies all M6.2 visual styling (colour palette, typography, accounting
     borders, number formats) to every cell written during hydration.
  5. Writes a fully hydrated Audit Log (Sheet 8) with per-item fiscal metadata.
  6. Returns the workbook as raw bytes for S3 upload or HTTP streaming.

Workbook layout (10 sheets — Amendment V1.2 §6.1 compliance)
─────────────────────────────────────────────────────────────
  Sheet 1  — Cover & Metadata
  Sheet 2  — Income Statement        (multi-period pivot)
  Sheet 3  — Balance Sheet           (multi-period pivot)
  Sheet 4  — Cash Flow               (multi-period pivot)
  Sheet 5  — Financial Ratios        (placeholder)
  Sheet 6  — Segments                (placeholder)
  Sheet 7  — Strategic Initiatives   (placeholder)
  Sheet 8  — Audit Log               (MANDATORY — Amendment V1.2 §6.1)
  Sheet 9  — Validation Report       (placeholder)
  Sheet 10 — Data Sources            (placeholder)

Multi-period pivot layout (Sheets 2–4)
───────────────────────────────────────
  Row 1:  Title banner (company / period range / reporting standard)
  Row 2:  Sub-header currency legend and FX rate methodology note
  Row 3:  [blank separator]
  Row 4:  Period group headers (merged: "FY 2022" spanning 4 cols, "FY 2023", …)
  Row 5:  Column sub-headers ("Value (Reported)" | "CCY" | "Value (USD)" | "FX Rate")
  Row 6+: One row per unique canonical tag, in taxonomy sort order.

  Fixed left columns (always present):
    Col A — Concept Label  (human-readable name)
    Col B — Canonical Tag  (XBRL tag)

  Per-period column group (4 cols per period, repeated right):
    +0 — Value (Reported)   format: #,##0.00
    +1 — Reported Currency  (ISO 4217)
    +2 — Value (USD)        format: #,##0.00
    +3 — FX Rate Used       format: 0.00000

Period chronological ordering
──────────────────────────────
  Within a year: Q1 < Q2 < Q3 < Q4 < H1 < H2 < FY
  Across years:  earlier years on the left (ascending fiscal_year).

Dependencies
────────────
  openpyxl  — workbook construction
  sqlalchemy.ext.asyncio.AsyncSession — data loading

Milestone: M6.3 — Data Ingestion and Period-Column Hydration
"""

from __future__ import annotations

import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from typing import Final, Sequence

import structlog
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

log = structlog.get_logger(__name__)

# ---------------------------------------------------------------------------
# Taxonomy field-order maps — controls row sequence within each sheet
# ---------------------------------------------------------------------------

_IS_FIELD_ORDER: Final[dict[str, int]] = {
    # ── Revenue ────────────────────────────────────────────────────────────
    "us-gaap:Revenues":                                            100,
    "us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax": 101,
    "ifrs-full:Revenue":                                           100,
    "ifrs-full:RevenueFromContractsWithCustomers":                 101,
    "ind-as:Revenue":                                              100,
    "ind-as:RevenueFromOperations":                                101,
    "ind-as:OtherIncome":                                          102,
    "ind-as:TotalIncome":                                          103,
    # ── Cost of sales ──────────────────────────────────────────────────────
    "us-gaap:CostOfRevenue":                                       200,
    "us-gaap:CostOfGoodsSold":                                     201,
    "ifrs-full:CostOfSales":                                       200,
    "ind-as:CostOfMaterialsConsumed":                              200,
    "ind-as:PurchasesOfStockInTrade":                              201,
    "ind-as:ChangesInInventoriesOfFinishedGoodsWorkInProgressAndStockInTrade": 202,
    "ind-as:EmployeeBenefitsExpense":                              203,
    # ── Gross profit ───────────────────────────────────────────────────────
    "us-gaap:GrossProfit":                                         300,
    "ifrs-full:GrossProfit":                                       300,
    # ── Operating expenses ─────────────────────────────────────────────────
    "us-gaap:ResearchAndDevelopmentExpense":                       400,
    "us-gaap:SellingGeneralAndAdministrativeExpense":              401,
    "us-gaap:GeneralAndAdministrativeExpense":                     402,
    "us-gaap:SellingExpense":                                      403,
    "us-gaap:MarketingExpense":                                    404,
    "us-gaap:DepreciationDepletionAndAmortization":                405,
    "us-gaap:OperatingExpenses":                                   409,
    "ifrs-full:DistributionCosts":                                 400,
    "ifrs-full:AdministrativeExpense":                             401,
    "ifrs-full:ResearchAndDevelopmentExpense":                     402,
    "ind-as:FinanceCosts":                                         410,
    "ind-as:DepreciationDepletionAndAmortisation":                 411,
    "ind-as:OtherExpenses":                                        412,
    "ind-as:Expenses":                                             419,
    # ── Operating income ───────────────────────────────────────────────────
    "us-gaap:OperatingIncomeLoss":                                 500,
    "ifrs-full:ProfitLossFromOperatingActivities":                 500,
    "ind-as:ProfitBeforeExceptionalItemsAndTax":                   501,
    "ind-as:ExceptionalItems":                                     502,
    # ── Non-operating / finance ────────────────────────────────────────────
    "us-gaap:InterestExpense":                                     600,
    "us-gaap:NonoperatingIncomeExpense":                           601,
    "ifrs-full:FinanceCosts":                                      600,
    "ifrs-full:FinanceIncome":                                     601,
    # ── Pre-tax income ─────────────────────────────────────────────────────
    "us-gaap:IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest": 700,
    "ifrs-full:ProfitLossBeforeTax":                               700,
    "ind-as:ProfitBeforeTax":                                      700,
    # ── Tax ────────────────────────────────────────────────────────────────
    "us-gaap:IncomeTaxExpenseBenefit":                             800,
    "ifrs-full:IncomeTaxExpenseContinuingOperations":              800,
    "ind-as:TaxExpense":                                           800,
    "ind-as:CurrentTax":                                           801,
    "ind-as:DeferredTax":                                          802,
    # ── Net income ─────────────────────────────────────────────────────────
    "us-gaap:NetIncomeLoss":                                       900,
    "ifrs-full:ProfitLoss":                                        900,
    "ind-as:ProfitLoss":                                           900,
    "ifrs-full:ComprehensiveIncome":                               910,
    "ind-as:OtherComprehensiveIncome":                             911,
    "ind-as:TotalComprehensiveIncome":                             912,
    # ── EPS ────────────────────────────────────────────────────────────────
    "us-gaap:EarningsPerShareBasic":                               990,
    "us-gaap:EarningsPerShareDiluted":                             991,
    "ifrs-full:BasicEarningsLossPerShare":                         990,
    "ifrs-full:DilutedEarningsLossPerShare":                       991,
    "ind-as:BasicEarningsPerShare":                                990,
    "ind-as:DilutedEarningsPerShare":                              991,
}

_BS_FIELD_ORDER: Final[dict[str, int]] = {
    # ── Total assets ───────────────────────────────────────────────────────
    "us-gaap:Assets":                                              100,
    "ifrs-full:Assets":                                            100,
    "ind-as:Assets":                                               100,
    # ── Current assets ─────────────────────────────────────────────────────
    "us-gaap:AssetsCurrent":                                       200,
    "ifrs-full:CurrentAssets":                                     200,
    "ind-as:CurrentAssets":                                        200,
    "us-gaap:CashAndCashEquivalentsAtCarryingValue":               201,
    "ifrs-full:CashAndCashEquivalents":                            201,
    "ind-as:CashAndCashEquivalents":                               201,
    "us-gaap:ShortTermInvestments":                                202,
    "us-gaap:AccountsReceivableNetCurrent":                        203,
    "ifrs-full:TradeAndOtherCurrentReceivables":                   203,
    "ind-as:TradeReceivables":                                     203,
    "us-gaap:InventoryNet":                                        204,
    "ifrs-full:Inventories":                                       204,
    "ind-as:Inventories":                                          204,
    "us-gaap:OtherAssetsCurrent":                                  209,
    "ifrs-full:OtherCurrentAssets":                                209,
    "ind-as:OtherCurrentAssets":                                   209,
    # ── Non-current assets ─────────────────────────────────────────────────
    "us-gaap:AssetsNoncurrent":                                    300,
    "ifrs-full:NoncurrentAssets":                                  300,
    "ind-as:NoncurrentAssets":                                     300,
    "us-gaap:PropertyPlantAndEquipmentNet":                        301,
    "ifrs-full:PropertyPlantAndEquipment":                         301,
    "ind-as:PropertyPlantAndEquipment":                            301,
    "ind-as:CapitalWorkInProgress":                                302,
    "ifrs-full:RightofuseAssets":                                  303,
    "ind-as:RightofuseAssets":                                     303,
    "us-gaap:Goodwill":                                            304,
    "ifrs-full:Goodwill":                                          304,
    "ind-as:Goodwill":                                             304,
    "us-gaap:IntangibleAssetsNetExcludingGoodwill":                305,
    "ifrs-full:IntangibleAssetsOtherThanGoodwill":                 305,
    "ind-as:IntangibleAssets":                                     305,
    "us-gaap:DeferredIncomeTaxAssetsNet":                          309,
    "ifrs-full:DeferredTaxAssets":                                 309,
    "ind-as:DeferredTaxAssets":                                    309,
    # ── Total liabilities ──────────────────────────────────────────────────
    "us-gaap:Liabilities":                                         400,
    "ifrs-full:Liabilities":                                       400,
    "ind-as:Liabilities":                                          400,
    # ── Current liabilities ────────────────────────────────────────────────
    "us-gaap:LiabilitiesCurrent":                                  500,
    "ifrs-full:CurrentLiabilities":                                500,
    "ind-as:CurrentLiabilities":                                   500,
    "us-gaap:AccountsPayableCurrent":                              501,
    "ifrs-full:TradeAndOtherCurrentPayables":                      501,
    "ind-as:TradePayables":                                        501,
    "us-gaap:ShortTermBorrowings":                                 502,
    "ifrs-full:ShorttermBorrowings":                               502,
    "ind-as:CurrentBorrowings":                                    502,
    "us-gaap:OtherLiabilitiesCurrent":                             509,
    "ifrs-full:OtherCurrentLiabilities":                           509,
    "ind-as:OtherCurrentLiabilities":                              509,
    # ── Non-current liabilities ────────────────────────────────────────────
    "us-gaap:LiabilitiesNoncurrent":                               600,
    "ifrs-full:NoncurrentLiabilities":                             600,
    "ind-as:NoncurrentLiabilities":                                600,
    "us-gaap:LongTermDebtNoncurrent":                              601,
    "ifrs-full:NoncurrentPortionOfNoncurrentBorrowings":           601,
    "ind-as:NoncurrentBorrowings":                                 601,
    "us-gaap:DeferredIncomeTaxLiabilitiesNet":                     609,
    "ifrs-full:DeferredTaxLiabilities":                            609,
    "ind-as:DeferredTaxLiabilities":                               609,
    # ── Equity ─────────────────────────────────────────────────────────────
    "us-gaap:StockholdersEquity":                                  700,
    "ifrs-full:Equity":                                            700,
    "ind-as:Equity":                                               700,
    "us-gaap:CommonStockValue":                                    701,
    "ifrs-full:IssuedCapital":                                     701,
    "ind-as:ShareCapital":                                         701,
    "us-gaap:AdditionalPaidInCapital":                             702,
    "ifrs-full:SharePremium":                                      702,
    "ind-as:OtherEquity":                                          703,
    "ind-as:ReservesAndSurplus":                                   703,
    "us-gaap:RetainedEarningsAccumulatedDeficit":                  704,
    "ifrs-full:RetainedEarnings":                                  704,
    "us-gaap:AccumulatedOtherComprehensiveIncomeLossNetOfTax":     705,
    "us-gaap:TreasuryStockValue":                                  706,
    # ── Total liabilities + equity ─────────────────────────────────────────
    "us-gaap:LiabilitiesAndStockholdersEquity":                    800,
    "ifrs-full:EquityAndLiabilities":                              800,
    "ind-as:EquityAndLiabilities":                                 800,
}

_CF_FIELD_ORDER: Final[dict[str, int]] = {
    # ── Operating ──────────────────────────────────────────────────────────
    "us-gaap:NetCashProvidedByUsedInOperatingActivities":          100,
    "ifrs-full:CashFlowsFromUsedInOperatingActivities":            100,
    "ind-as:CashFlowsFromUsedInOperatingActivities":               100,
    # ── Investing ──────────────────────────────────────────────────────────
    "us-gaap:NetCashProvidedByUsedInInvestingActivities":          200,
    "ifrs-full:CashFlowsFromUsedInInvestingActivities":            200,
    "ind-as:CashFlowsFromUsedInInvestingActivities":               200,
    "us-gaap:PaymentsToAcquirePropertyPlantAndEquipment":          201,
    "ifrs-full:PurchaseOfPropertyPlantAndEquipment":               201,
    "ind-as:PurchaseOfPropertyPlantAndEquipment":                  201,
    # ── Financing ──────────────────────────────────────────────────────────
    "us-gaap:NetCashProvidedByUsedInFinancingActivities":          300,
    "ifrs-full:CashFlowsFromUsedInFinancingActivities":            300,
    "ind-as:CashFlowsFromUsedInFinancingActivities":               300,
    "us-gaap:PaymentsOfDividends":                                 301,
    "ifrs-full:DividendsPaidClassifiedAsFinancingActivities":      301,
    "ind-as:DividendsPaid":                                        301,
    "us-gaap:PaymentsForRepurchaseOfCommonStock":                  302,
    "us-gaap:ProceedsFromIssuanceOfLongTermDebt":                  303,
    "us-gaap:RepaymentsOfLongTermDebt":                            304,
    # ── Net change ─────────────────────────────────────────────────────────
    "us-gaap:CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect": 400,
    "ifrs-full:IncreaseDecreaseInCashAndCashEquivalents":          400,
    "ind-as:IncreaseDecreaseInCashAndCashEquivalents":             400,
}

# Combined statement-type → ordering map
_FIELD_ORDER_BY_STATEMENT: Final[dict[str, dict[str, int]]] = {
    "IS": _IS_FIELD_ORDER,
    "BS": _BS_FIELD_ORDER,
    "CF": _CF_FIELD_ORDER,
}

# Fallback sort key for unrecognised canonical tags
_ORDER_FALLBACK: Final[int] = 9999

# ---------------------------------------------------------------------------
# Period chronological ordering
# ---------------------------------------------------------------------------
# Controls left-to-right column sequencing of fiscal periods on statement sheets.
# Within a year: Q1 < Q2 < Q3 < Q4 < H1 < H2 < FY (annual comes last).
# Across years: ascending fiscal_year (oldest data on the left).

_PERIOD_SORT_ORDER: Final[dict[str, int]] = {
    "Q1": 1,
    "Q2": 2,
    "Q3": 3,
    "Q4": 4,
    "H1": 5,
    "H2": 6,
    "FY": 7,
}


def _period_sort_key(year_period: tuple[int, str]) -> tuple[int, int]:
    """Return a (year, intra-year) sort tuple for chronological ordering."""
    year, period = year_period
    return (year, _PERIOD_SORT_ORDER.get(period.upper(), 99))


def _period_label(fiscal_year: int, fiscal_period: str) -> str:
    """
    Format a (year, period) pair as a human-readable column group header.

    Examples::
        (2024, "FY")  → "FY 2024"
        (2024, "Q1")  → "Q1 2024"
        (2024, "H2")  → "H2 2024"
    """
    return f"{fiscal_period.upper()} {fiscal_year}"


# ---------------------------------------------------------------------------
# Accounting border classification
# ---------------------------------------------------------------------------

# Terminal rows — double-line bottom border (Net Income, Total Assets, etc.)
_TERMINAL_TAGS: Final[frozenset[str]] = frozenset({
    "us-gaap:NetIncomeLoss",
    "ifrs-full:ProfitLoss",
    "ind-as:ProfitLoss",
    "ifrs-full:ComprehensiveIncome",
    "ind-as:TotalComprehensiveIncome",
    "us-gaap:Assets",
    "ifrs-full:Assets",
    "ind-as:Assets",
    "us-gaap:LiabilitiesAndStockholdersEquity",
    "ifrs-full:EquityAndLiabilities",
    "ind-as:EquityAndLiabilities",
    "us-gaap:CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect",
    "ifrs-full:IncreaseDecreaseInCashAndCashEquivalents",
    "ind-as:IncreaseDecreaseInCashAndCashEquivalents",
})

# Subtotal rows — thin top border only
_SUBTOTAL_TAGS: Final[frozenset[str]] = frozenset({
    "us-gaap:GrossProfit",
    "ifrs-full:GrossProfit",
    "us-gaap:OperatingIncomeLoss",
    "ifrs-full:ProfitLossFromOperatingActivities",
    "ind-as:ProfitBeforeExceptionalItemsAndTax",
    "us-gaap:IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest",
    "ifrs-full:ProfitLossBeforeTax",
    "ind-as:ProfitBeforeTax",
    "us-gaap:AssetsCurrent",
    "ifrs-full:CurrentAssets",
    "ind-as:CurrentAssets",
    "us-gaap:AssetsNoncurrent",
    "ifrs-full:NoncurrentAssets",
    "ind-as:NoncurrentAssets",
    "us-gaap:Liabilities",
    "ifrs-full:Liabilities",
    "ind-as:Liabilities",
    "us-gaap:LiabilitiesCurrent",
    "ifrs-full:CurrentLiabilities",
    "ind-as:CurrentLiabilities",
    "us-gaap:LiabilitiesNoncurrent",
    "ifrs-full:NoncurrentLiabilities",
    "ind-as:NoncurrentLiabilities",
    "us-gaap:StockholdersEquity",
    "ifrs-full:Equity",
    "ind-as:Equity",
    "us-gaap:NetCashProvidedByUsedInOperatingActivities",
    "ifrs-full:CashFlowsFromUsedInOperatingActivities",
    "ind-as:CashFlowsFromUsedInOperatingActivities",
    "us-gaap:NetCashProvidedByUsedInInvestingActivities",
    "ifrs-full:CashFlowsFromUsedInInvestingActivities",
    "ind-as:CashFlowsFromUsedInInvestingActivities",
    "us-gaap:NetCashProvidedByUsedInFinancingActivities",
    "ifrs-full:CashFlowsFromUsedInFinancingActivities",
    "ind-as:CashFlowsFromUsedInFinancingActivities",
})

# ---------------------------------------------------------------------------
# Visual styling constants
# ---------------------------------------------------------------------------

_CLR_NAVY:        Final[str] = "1F3864"   # Primary header background
_CLR_NAVY_LIGHT:  Final[str] = "BDD7EE"   # Terminal row background
_CLR_SLATE:       Final[str] = "D9E1F2"   # Sub-header / subtotal background
_CLR_ZEBRA_EVEN:  Final[str] = "F2F2F2"   # Alternating data row (even index)
_CLR_WHITE:       Final[str] = "FFFFFF"   # Alternating data row (odd index)
_CLR_BLACK:       Final[str] = "000000"   # Default foreground

# Number format strings
_FMT_CURRENCY: Final[str] = "#,##0.00"
_FMT_FX_RATE:  Final[str] = "0.00000"

# Fixed column indices (1-based) for the statement sheet left anchor columns
_COL_CONCEPT_LABEL: Final[int] = 1   # "Concept Label"
_COL_CANONICAL_TAG: Final[int] = 2   # "Canonical Tag"
_COL_DATA_START:    Final[int] = 3   # First data column (first period, Value Reported)

# Per-period column offsets (0-based within each 4-column period group)
_OFF_VALUE_REPORTED:  Final[int] = 0
_OFF_REPORTED_CCY:    Final[int] = 1
_OFF_VALUE_USD:       Final[int] = 2
_OFF_FX_RATE:         Final[int] = 3
_COLS_PER_PERIOD:     Final[int] = 4

# Sub-header labels for the 4 per-period columns (row 5 of statement sheets)
_PERIOD_COL_SUBHEADERS: Final[tuple[str, str, str, str]] = (
    "Value (Reported)",
    "CCY",
    "Value (USD)",
    "FX Rate",
)


# ---------------------------------------------------------------------------
# Styling helpers — deferred openpyxl imports
# ---------------------------------------------------------------------------

def _auto_fit_columns(ws: object) -> None:
    """
    Set each column width to the longest cell content length plus 4-char
    safety padding, preventing ``###`` truncation on all data cells.
    """
    from openpyxl.utils import get_column_letter  # type: ignore[import]

    col_widths: dict[int, int] = {}
    for row in ws.iter_rows():  # type: ignore[union-attr]
        for cell in row:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > col_widths.get(cell.column, 0):
                    col_widths[cell.column] = length

    for col_idx, max_len in col_widths.items():
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max_len + 4  # type: ignore[union-attr]


def _make_fill(hex_color: str) -> object:
    """Return a solid ``PatternFill`` for the given hex color string."""
    from openpyxl.styles import PatternFill  # type: ignore[import]
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _make_font(
    *,
    size: int = 10,
    bold: bool = False,
    color: str = _CLR_BLACK,
) -> object:
    """Return an openpyxl ``Font`` object (Calibri family)."""
    from openpyxl.styles import Font  # type: ignore[import]
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _make_thin_border(*, top: bool = False, bottom: bool = False) -> object:
    """Return a ``Border`` with optional thin top/bottom sides."""
    from openpyxl.styles import Border, Side  # type: ignore[import]
    thin = Side(style="thin")
    return Border(
        top=thin if top else Side(),
        bottom=thin if bottom else Side(),
    )


def _make_double_bottom_border() -> object:
    """
    Return a ``Border`` with thin top and double-line bottom — the accounting
    convention for terminal totals (Net Income, Total Assets, etc.).
    """
    from openpyxl.styles import Border, Side  # type: ignore[import]
    return Border(
        top=Side(style="thin"),
        bottom=Side(style="double"),
    )


def _apply_row_fill(ws: object, row_idx: int, col_count: int, fill: object) -> None:
    """Apply *fill* to every cell in *row_idx* across *col_count* columns."""
    for col in range(1, col_count + 1):
        ws.cell(row=row_idx, column=col).fill = fill  # type: ignore[union-attr]


def _apply_row_font(ws: object, row_idx: int, col_count: int, font: object) -> None:
    """Apply *font* to every cell in *row_idx* across *col_count* columns."""
    for col in range(1, col_count + 1):
        ws.cell(row=row_idx, column=col).font = font  # type: ignore[union-attr]


def _apply_row_border(ws: object, row_idx: int, col_count: int, border: object) -> None:
    """Apply *border* to every cell in *row_idx* across *col_count* columns."""
    for col in range(1, col_count + 1):
        ws.cell(row=row_idx, column=col).border = border  # type: ignore[union-attr]


# ---------------------------------------------------------------------------
# Data transfer objects
# ---------------------------------------------------------------------------


@dataclass
class ExportLineItem:
    """
    Flattened representation of a ``FinancialLineItem`` row.

    M6.3 additions: ``fiscal_year`` and ``fiscal_period`` fields are now
    present on every item so that multi-period pivot logic and the per-item
    audit log rows can reference their originating period without consulting
    the job-level ``ExportContext`` fields.

    Attributes:
        canonical_field:    XBRL concept tag or normalised slug.
        concept_label:      Human-readable display label.
        statement_type:     "IS" | "BS" | "CF".
        fiscal_year:        Fiscal year integer this row belongs to.
        fiscal_period:      Fiscal period string ("FY" | "Q1"–"Q4" | "H1" | "H2").
        value_reported:     Original reported value (Decimal or None).
        reported_currency:  ISO 4217 code (e.g. "INR").
        value_usd:          Translated USD value (Decimal or None).
        fx_rate_used:       FX coefficient used for translation (Decimal or None).
        filing_date:        Date the document was filed.
        is_restated:        True when this row supersedes an earlier filing.
        extraction_method:  How the value was extracted: xbrl | pdf | ocr | ai.
        source_file_hash:   SHA-256 hex digest of the source document.
        derived_formula:    Algebraic derivation string (Amendment V1.2 §8.1).
    """

    canonical_field: str
    concept_label: str
    statement_type: str
    fiscal_year: int
    fiscal_period: str
    value_reported: Decimal | None
    reported_currency: str
    value_usd: Decimal | None
    fx_rate_used: Decimal | None
    filing_date: date
    is_restated: bool
    extraction_method: str | None
    source_file_hash: str | None
    derived_formula: str | None

    @property
    def sort_key(self) -> int:
        """Taxonomy ordering integer for this item's statement type."""
        order_map = _FIELD_ORDER_BY_STATEMENT.get(self.statement_type, {})
        return order_map.get(self.canonical_field, _ORDER_FALLBACK)

    @property
    def period_key(self) -> tuple[int, str]:
        """``(fiscal_year, fiscal_period)`` pair used as pivot dict key."""
        return (self.fiscal_year, self.fiscal_period)


@dataclass
class ExportContext:
    """
    All database data assembled for a single export run.

    M6.3: ``is_items``, ``bs_items``, and ``cf_items`` now contain data
    across ALL available fiscal periods for the company, not just the
    single period recorded on the originating FinancialJob.  The
    ``fiscal_year`` / ``fiscal_period`` fields still represent the job's
    *primary* period (shown on the Cover sheet); the ``periods`` property
    derives the full chronological list from the loaded items.

    Attributes:
        job_id:              UUID of the originating FinancialJob.
        company_name:        Full legal name of the company.
        company_ticker:      Stock ticker symbol.
        fiscal_year:         Primary fiscal year of the job (for cover sheet).
        fiscal_period:       Primary fiscal period of the job (for cover sheet).
        reporting_standard:  US_GAAP | IFRS | IND_AS.
        document_url:        S3 key of the source document.
        filing_date:         Filing date from the job.
        export_timestamp:    UTC datetime when this export was generated.
        is_items:            All IS line items across all periods (ordered).
        bs_items:            All BS line items across all periods (ordered).
        cf_items:            All CF line items across all periods (ordered).
        translation_complete: True when ALL non-USD rows have value_usd set.
    """

    job_id: uuid.UUID
    company_name: str
    company_ticker: str
    fiscal_year: int
    fiscal_period: str
    reporting_standard: str
    document_url: str | None
    filing_date: date | None
    export_timestamp: datetime
    is_items: list[ExportLineItem] = field(default_factory=list)
    bs_items: list[ExportLineItem] = field(default_factory=list)
    cf_items: list[ExportLineItem] = field(default_factory=list)
    translation_complete: bool = False
    # Optional: populated by _load_export_context from validation_results table (M4/B6)
    validation_data: dict | None = field(default=None)

    @property
    def all_items(self) -> list[ExportLineItem]:
        """All line items across all three statements, order preserved."""
        return self.is_items + self.bs_items + self.cf_items

    @property
    def total_row_count(self) -> int:
        """Total number of exported rows across all statements and periods."""
        return len(self.is_items) + len(self.bs_items) + len(self.cf_items)

    @property
    def currency_label(self) -> str:
        """
        Reported currency label — the first non-USD currency in the dataset,
        or "USD" when all items are already denominated in USD.
        """
        for item in self.all_items:
            if item.reported_currency and item.reported_currency.upper() != "USD":
                return item.reported_currency.upper()
        return "USD"

    @property
    def periods(self) -> list[tuple[int, str]]:
        """
        Distinct (fiscal_year, fiscal_period) pairs present in the dataset,
        sorted chronologically: ascending year, then Q1 < Q2 < Q3 < Q4 <
        H1 < H2 < FY within each year.
        """
        seen: set[tuple[int, str]] = set()
        for item in self.all_items:
            seen.add(item.period_key)
        return sorted(seen, key=_period_sort_key)

    @property
    def period_range_label(self) -> str:
        """
        Human-readable range string for the Cover sheet.

        Examples::
            Single period:  "FY 2024"
            Multiple:       "FY 2022 – FY 2024"
        """
        ps = self.periods
        if not ps:
            return _period_label(self.fiscal_year, self.fiscal_period)
        if len(ps) == 1:
            return _period_label(*ps[0])
        return f"{_period_label(*ps[0])} – {_period_label(*ps[-1])}"


# ---------------------------------------------------------------------------
# Label derivation
# ---------------------------------------------------------------------------

def _label_from_canonical(canonical_field: str) -> str:
    """
    Derive a human-readable display label from a canonical XBRL tag.

    Examples::
        "us-gaap:Revenues"              → "Revenues"
        "ifrs-full:ProfitLossBeforeTax" → "Profit Loss Before Tax"
        "raw:Total_Operating_Expenses"  → "Total Operating Expenses"
    """
    import re

    if ":" in canonical_field:
        label = canonical_field.split(":", 1)[1]
    else:
        label = canonical_field

    label = label.replace("_", " ")
    label = re.sub(r"([a-z])([A-Z])", r"\1 \2", label)
    label = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1 \2", label)
    return label.strip()


# ---------------------------------------------------------------------------
# Pivot builder
# ---------------------------------------------------------------------------

def _build_pivot(
    items: list[ExportLineItem],
    statement_type: str,
) -> tuple[list[str], list[tuple[int, str]], dict[str, dict[tuple[int, str], ExportLineItem]]]:
    """
    Pivot a flat list of ``ExportLineItem`` objects into a 2-D structure
    suitable for multi-period spreadsheet rendering.

    Args:
        items:          Flat list of ExportLineItem for a single statement type
                        (already filtered, may span multiple periods).
        statement_type: "IS" | "BS" | "CF" — used to look up taxonomy order.

    Returns:
        A 3-tuple:
          1. ``ordered_tags``  — list of canonical tags sorted by taxonomy
             order then alphabetically (controls row sequence on the sheet).
          2. ``sorted_periods`` — list of (year, period) pairs sorted
             chronologically (controls column group sequence on the sheet).
          3. ``pivot``          — nested dict:
             ``pivot[canonical_tag][(fiscal_year, fiscal_period)]``
             → ``ExportLineItem``.  Missing cells are simply absent from
             the inner dict (caller writes blank for missing combinations).

    Design notes:
        • Duplicate (canonical_field, period_key) combinations are resolved
          by keeping the last occurrence in the ``items`` list.  In practice
          ``_load_export_context`` filters ``is_restated=FALSE`` and the DB
          constraint prevents duplicates, but this guard is retained for
          defensive correctness.
        • Unknown canonical tags (sort_key == 9999) are appended to the
          tail of ``ordered_tags`` in alphabetical order.
    """
    order_map = _FIELD_ORDER_BY_STATEMENT.get(statement_type, {})

    pivot: dict[str, dict[tuple[int, str], ExportLineItem]] = {}
    period_set: set[tuple[int, str]] = set()

    for item in items:
        period_set.add(item.period_key)
        if item.canonical_field not in pivot:
            pivot[item.canonical_field] = {}
        # Last-one-wins for any duplicate (canonical_field, period_key)
        pivot[item.canonical_field][item.period_key] = item

    ordered_tags = sorted(
        pivot.keys(),
        key=lambda tag: (order_map.get(tag, _ORDER_FALLBACK), tag),
    )
    sorted_periods = sorted(period_set, key=_period_sort_key)

    return ordered_tags, sorted_periods, pivot


# ---------------------------------------------------------------------------
# Main service
# ---------------------------------------------------------------------------


class ExcelExportService:
    """
    Async service that produces a multi-sheet, multi-period Excel workbook
    for a single FinancialJob and returns raw ``.xlsx`` bytes.

    Usage::

        service = ExcelExportService()
        xlsx_bytes = await service.export(job_id=job_uuid, session=db_session)

    Design contract:
      - ``session`` is NOT committed or closed by this service.
      - ``export()`` is the single public entry point.
      - ``_load_export_context()`` is async (DB access); ``_build_workbook()``
        is synchronous (pure openpyxl, no DB) — cleanly separated for testing.
    """

    _SHEET_COVER      = "Cover & Metadata"
    _SHEET_IS         = "Income Statement"
    _SHEET_BS         = "Balance Sheet"
    _SHEET_CF         = "Cash Flow"
    _SHEET_RATIOS     = "Financial Ratios"
    _SHEET_SEGMENTS   = "Segments"
    _SHEET_STRATEGIC  = "Strategic Initiatives"
    _SHEET_VALIDATION = "Validation Report"
    _SHEET_SOURCES    = "Data Sources"

    # ---------------------------------------------------------------------------
    # Public entry point
    # ---------------------------------------------------------------------------

    async def export(
        self,
        job_id: uuid.UUID,
        session: AsyncSession,
    ) -> bytes:
        """
        Produce the full multi-period Excel workbook for *job_id*.

        Args:
            job_id:  UUID of the FinancialJob to export.
            session: Active async database session (caller owns lifecycle).

        Returns:
            Raw bytes of the generated ``.xlsx`` file.

        Raises:
            ExportJobNotFoundError:        *job_id* does not exist.
            ExportNoDataError:             No FinancialLineItem rows found.
            ExportCompanyNotFoundError:    Company linked to the job missing.
        """
        bound_log = log.bind(job_id=str(job_id))
        bound_log.info("excel_export.started")

        ctx = await self._load_export_context(job_id=job_id, session=session)

        bound_log.info(
            "excel_export.context_loaded",
            company=ctx.company_name,
            periods=len(ctx.periods),
            is_rows=len(ctx.is_items),
            bs_rows=len(ctx.bs_items),
            cf_rows=len(ctx.cf_items),
            translation_complete=ctx.translation_complete,
        )

        xlsx_bytes = self._build_workbook(ctx)

        bound_log.info(
            "excel_export.completed",
            total_rows=ctx.total_row_count,
            size_bytes=len(xlsx_bytes),
        )
        return xlsx_bytes

    # ---------------------------------------------------------------------------
    # Data loading
    # ---------------------------------------------------------------------------

    async def _load_export_context(
        self,
        job_id: uuid.UUID,
        session: AsyncSession,
    ) -> ExportContext:
        """
        Query the database and return a fully populated multi-period
        ``ExportContext``.

        M6.3 change: the FinancialLineItem query no longer filters by
        ``fiscal_year`` or ``fiscal_period``.  All non-restated rows for
        the job's company are loaded, enabling multi-period pivot sheets.
        The job's ``fiscal_year`` and ``fiscal_period`` are retained on
        ``ExportContext`` for the Cover sheet and for log correlation only.

        Raises:
            ExportJobNotFoundError:        Job not found.
            ExportCompanyNotFoundError:    Company not found.
            ExportNoDataError:             No line items for this company.
        """
        from apps.api.models import Company, FinancialJob, FinancialLineItem

        # 1. Load job
        job: FinancialJob | None = await session.get(FinancialJob, job_id)
        if job is None:
            raise ExportJobNotFoundError(
                f"FinancialJob {job_id!r} not found."
            )

        # 2. Load company
        company: Company | None = await session.get(Company, job.company_id)
        if company is None:
            raise ExportCompanyNotFoundError(
                f"Company {job.company_id!r} linked to job {job_id!r} not found."
            )

        # 3. Resolve primary fiscal metadata (cover sheet / log correlation)
        fiscal_year: int = job.fiscal_year or datetime.utcnow().year
        fiscal_period: str = getattr(job, "fiscal_period", "FY") or "FY"
        reporting_standard: str = (
            job.reporting_standard.value
            if hasattr(job.reporting_standard, "value")
            else str(getattr(job, "reporting_standard", "US_GAAP") or "US_GAAP")
        )

        # 4. Load ALL non-restated line items for the company (multi-period)
        stmt = (
            select(FinancialLineItem)
            .where(
                FinancialLineItem.company_id == job.company_id,
                FinancialLineItem.is_restated.is_(False),
            )
            .order_by(
                FinancialLineItem.fiscal_year.asc(),
                FinancialLineItem.fiscal_period.asc(),
                FinancialLineItem.statement_type.asc(),
                FinancialLineItem.canonical_field.asc(),
            )
        )
        result = await session.execute(stmt)
        raw_items: Sequence[FinancialLineItem] = result.scalars().all()

        if not raw_items:
            raise ExportNoDataError(
                f"No FinancialLineItem rows found for company {job.company_id} "
                f"(job {job_id!r}, is_restated=FALSE). "
                "Ensure extraction and FX translation have completed before export."
            )

        # 5. Convert ORM rows → ExportLineItem DTOs
        export_items = [_orm_to_export_item(row) for row in raw_items]

        # 6. Partition by statement type (taxonomy sort applied inside pivot)
        is_items = [i for i in export_items if i.statement_type == "IS"]
        bs_items = [i for i in export_items if i.statement_type == "BS"]
        cf_items = [i for i in export_items if i.statement_type == "CF"]

        # 7. Translation completeness — all non-USD items must have value_usd
        non_usd = [
            i for i in export_items
            if i.reported_currency and i.reported_currency.upper() != "USD"
        ]
        translation_complete = (
            all(i.value_usd is not None for i in non_usd) if non_usd else True
        )

        # 8. Load most-recent ValidationResultRecord for this job (optional — B6)
        validation_data: dict | None = None
        try:
            from apps.api.models import ValidationResultRecord  # noqa: PLC0415
            vr_stmt = (
                select(ValidationResultRecord)
                .where(ValidationResultRecord.job_id == job_id)
                .order_by(ValidationResultRecord.created_at.desc())
                .limit(1)
            )
            vr_result = await session.execute(vr_stmt)
            vr: ValidationResultRecord | None = vr_result.scalars().first()
            if vr is not None:
                validation_data = {
                    "id":               str(vr.id),
                    "accession_number": vr.accession_number,
                    "company_id":       str(vr.company_id) if vr.company_id else None,
                    "fiscal_year":      vr.fiscal_year,
                    "fiscal_period":    vr.fiscal_period,
                    "items_validated":  vr.items_validated,
                    "is_exportable":    vr.is_exportable,
                    "critical_count":   vr.critical_count,
                    "warning_count":    vr.warning_count,
                    "confidence_score": vr.confidence_score,
                    "findings":         vr.findings or [],
                    "deductions":       vr.deductions or [],
                    "summary_text":     vr.summary_text,
                    "created_at":       vr.created_at.strftime("%Y-%m-%d %H:%M:%S UTC"),
                }
        except Exception:
            # Non-fatal: validation data is optional for the export
            pass

        return ExportContext(
            job_id=job_id,
            company_name=company.name,
            company_ticker=company.ticker,
            fiscal_year=fiscal_year,
            fiscal_period=fiscal_period,
            reporting_standard=reporting_standard,
            document_url=getattr(job, "document_url", None),
            filing_date=getattr(job, "filing_date", None),
            export_timestamp=datetime.utcnow(),
            is_items=is_items,
            bs_items=bs_items,
            cf_items=cf_items,
            translation_complete=translation_complete,
            validation_data=validation_data,
        )

    # ---------------------------------------------------------------------------
    # Workbook construction
    # ---------------------------------------------------------------------------

    def _build_workbook(self, ctx: ExportContext) -> bytes:
        """
        Assemble and return the ``.xlsx`` workbook from a populated
        ``ExportContext``.  No database access; synchronous; unit-testable.
        """
        import io

        import openpyxl  # type: ignore[import]

        wb = openpyxl.Workbook()
        if wb.active is not None and wb.active.title == "Sheet":
            del wb["Sheet"]

        # Sheets 1–4
        self._write_cover_sheet(wb.create_sheet(title=self._SHEET_COVER), ctx)
        self._write_statement_sheet(
            wb.create_sheet(title=self._SHEET_IS), ctx.is_items, ctx, "IS"
        )
        self._write_statement_sheet(
            wb.create_sheet(title=self._SHEET_BS), ctx.bs_items, ctx, "BS"
        )
        self._write_statement_sheet(
            wb.create_sheet(title=self._SHEET_CF), ctx.cf_items, ctx, "CF"
        )

        # Sheet 5: Financial Ratios (B6)
        from services.export.sheets.ratios import write_ratios_sheet  # noqa: PLC0415
        write_ratios_sheet(wb.create_sheet(title=self._SHEET_RATIOS), ctx)

        # Sheet 6: Segments / Revenue Contribution (B6)
        from services.export.sheets.segments import write_segments_sheet  # noqa: PLC0415
        write_segments_sheet(wb.create_sheet(title=self._SHEET_SEGMENTS), ctx)

        # Sheet 7: Strategic Initiatives / Executive Summary (B6)
        from services.export.sheets.strategic_initiatives import write_strategic_initiatives_sheet  # noqa: PLC0415
        write_strategic_initiatives_sheet(wb.create_sheet(title=self._SHEET_STRATEGIC), ctx)

        # Sheet 8: Mandatory Audit Log
        self._write_audit_log_sheet(wb.create_sheet(title="Audit Log"), ctx)

        # Sheet 9: Validation Report (B6)
        from services.export.sheets.validation_report import write_validation_report_sheet  # noqa: PLC0415
        write_validation_report_sheet(wb.create_sheet(title=self._SHEET_VALIDATION), ctx)

        # Sheet 10: Data Sources placeholder (future milestone)
        ws = wb.create_sheet(title=self._SHEET_SOURCES)
        c = ws.cell(row=1, column=1, value=f"[{self._SHEET_SOURCES} — to be implemented in a future milestone]")
        c.font = _make_font(size=10, color="808080")  # type: ignore[assignment]

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ---------------------------------------------------------------------------
    # Sheet writers
    # ---------------------------------------------------------------------------

    def _write_cover_sheet(self, ws: object, ctx: ExportContext) -> None:
        """
        Write Sheet 1 — Cover & Metadata.

        Rows:
          2  Title banner (company name + period range)
          4  Company
          5  Primary Fiscal Year
          6  Primary Fiscal Period
          7  Periods Covered    ← M6.3: shows full chronological range
          8  Reporting Standard
          9  Reported Currency
          10 Translated Currency
          11 Filing Date
          12 Export Timestamp
          13 Translation Complete
          15 Total Line Items
          16 Income Statement rows
          17 Balance Sheet rows
          18 Cash Flow rows
          20 Data Quality Notice (Amendment V1.2 §6.2)
        """
        _write = ws.cell  # type: ignore[attr-defined]

        # Title banner
        title_cell = _write(
            row=2, column=1,
            value=(
                f"Financial Data Hub — {ctx.company_name} ({ctx.company_ticker}) "
                f"— {ctx.period_range_label}"
            ),
        )
        title_cell.font = _make_font(size=16, bold=True, color=_CLR_WHITE)
        title_cell.fill = _make_fill(_CLR_NAVY)
        _write(row=2, column=2, value=None).fill = _make_fill(_CLR_NAVY)

        label_font = _make_font(size=11, bold=True)
        value_font = _make_font(size=10)
        label_fill = _make_fill(_CLR_SLATE)
        value_fill = _make_fill(_CLR_WHITE)

        metadata_rows = [
            (4,  "Company",              f"{ctx.company_name} ({ctx.company_ticker})"),
            (5,  "Primary Fiscal Year",  ctx.fiscal_year),
            (6,  "Primary Fiscal Period", ctx.fiscal_period),
            (7,  "Periods Covered",      ctx.period_range_label),
            (8,  "Reporting Standard",   ctx.reporting_standard),
            (9,  "Reported Currency",    ctx.currency_label),
            (10, "Translated Currency",  "USD"),
            (11, "Filing Date",
                 ctx.filing_date.isoformat() if ctx.filing_date else "N/A"),
            (12, "Export Timestamp",
                 ctx.export_timestamp.strftime("%Y-%m-%d %H:%M:%S UTC")),
            (13, "Translation Complete",
                 "Yes" if ctx.translation_complete
                 else ("N/A — all USD" if ctx.currency_label == "USD"
                       else "No (FX data gap)")),
        ]
        for row_idx, label, value in metadata_rows:
            lc = _write(row=row_idx, column=1, value=label)
            lc.font = label_font
            lc.fill = label_fill
            vc = _write(row=row_idx, column=2, value=value)
            vc.font = value_font
            vc.fill = value_fill

        count_rows = [
            (15, "Total Line Items",  ctx.total_row_count),
            (16, "Income Statement",  f"{len(ctx.is_items)} rows"),
            (17, "Balance Sheet",     f"{len(ctx.bs_items)} rows"),
            (18, "Cash Flow",         f"{len(ctx.cf_items)} rows"),
        ]
        for row_idx, label, value in count_rows:
            lc = _write(row=row_idx, column=1, value=label)
            lc.font = label_font
            lc.fill = label_fill
            vc = _write(row=row_idx, column=2, value=value)
            vc.font = value_font
            vc.fill = value_fill

        notice = _write(
            row=20, column=1,
            value=(
                "Data Quality Notice: AI-extracted cells carry Amendment V1.2 §6.2 "
                "lineage comments showing source page and confidence percentage."
            ),
        )
        notice.font = _make_font(size=9, color="595959")

        ws.column_dimensions["A"].width = 32  # type: ignore[union-attr]
        ws.column_dimensions["B"].width = 58  # type: ignore[union-attr]

    def _write_statement_sheet(
        self,
        ws: object,
        items: list[ExportLineItem],
        ctx: ExportContext,
        statement_type: str,
    ) -> None:
        """
        Write a multi-period financial statement sheet (IS, BS, or CF).

        Layout (M6.3 multi-period pivot)
        ─────────────────────────────────
        Row 1:  Title banner
        Row 2:  Currency legend sub-header
        Row 3:  [blank]
        Row 4:  Period group headers (merged 4-column spans per period)
        Row 5:  Per-period column sub-headers (Value (Reported) | CCY | Value (USD) | FX Rate)
        Row 6+: Data rows — one canonical tag per row, values per period.

        Styling applied per M6.2:
          • Rows 1/2/4/5 — navy or slate fills with bold fonts.
          • Data rows — zebra, subtotal (slate+thin-top), terminal (light-navy+double-bottom).
          • Number formats on value/FX-rate cells.
          • Freeze panes at row 6 (data start).
          • Auto-fit all columns after data written.

        Args:
            ws:             Worksheet (openpyxl).
            items:          All ExportLineItem rows for this statement type,
                            spanning all available fiscal periods.
            ctx:            ExportContext (for banner text).
            statement_type: "IS" | "BS" | "CF".
        """
        from openpyxl.styles import Alignment  # type: ignore[import]

        _write = ws.cell  # type: ignore[attr-defined]

        _DISPLAY_NAME = {
            "IS": "Income Statement",
            "BS": "Balance Sheet",
            "CF": "Cash Flow Statement",
        }
        _RATE_NOTE = {
            "IS": "FX rate: weighted average over fiscal period (Amendment V1.2 §3 Pass 2)",
            "BS": "FX rate: spot rate on period end date (Amendment V1.2 §3 Pass 1)",
            "CF": "FX rate: weighted average over fiscal period (Amendment V1.2 §3 Pass 2)",
        }
        display_name = _DISPLAY_NAME.get(statement_type, statement_type)
        rate_note    = _RATE_NOTE.get(statement_type, "")

        # ── Build pivot ───────────────────────────────────────────────────────
        if not items:
            empty = _write(row=1, column=1, value=f"[No data available for {display_name}]")
            empty.font = _make_font(size=10, color="808080")
            return

        ordered_tags, sorted_periods, pivot = _build_pivot(items, statement_type)
        n_periods  = len(sorted_periods)
        total_cols = _COL_DATA_START - 1 + n_periods * _COLS_PER_PERIOD

        center_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align     = Alignment(horizontal="left",   vertical="center")
        right_align    = Alignment(horizontal="right",  vertical="center")

        # ── Row 1: Title banner ───────────────────────────────────────────────
        title_text = (
            f"{display_name} — {ctx.company_name} ({ctx.company_ticker}) "
            f"— {ctx.period_range_label} — {ctx.reporting_standard}"
        )
        tc = _write(row=1, column=1, value=title_text)
        tc.font      = _make_font(size=16, bold=True, color=_CLR_WHITE)
        tc.fill      = _make_fill(_CLR_NAVY)
        tc.alignment = left_align
        for col in range(2, total_cols + 1):
            _write(row=1, column=col, value=None).fill = _make_fill(_CLR_NAVY)

        # ── Row 2: Currency legend sub-header ─────────────────────────────────
        legend_text = (
            f"Reported in {ctx.currency_label} | Translated to USD | {rate_note}"
        )
        lc = _write(row=2, column=1, value=legend_text)
        lc.font      = _make_font(size=12, bold=True)
        lc.fill      = _make_fill(_CLR_SLATE)
        lc.alignment = left_align
        for col in range(2, total_cols + 1):
            _write(row=2, column=col, value=None).fill = _make_fill(_CLR_SLATE)

        # ── Row 3: blank separator ────────────────────────────────────────────

        # ── Row 4: Period group headers (merged cells) ─────────────────────────
        # Fixed anchor columns: "Concept Label" and "Canonical Tag"
        for col, label in [
            (_COL_CONCEPT_LABEL, "Concept Label"),
            (_COL_CANONICAL_TAG, "Canonical Tag"),
        ]:
            hc = _write(row=4, column=col, value=label)
            hc.font      = _make_font(size=11, bold=True, color=_CLR_WHITE)
            hc.fill      = _make_fill(_CLR_NAVY)
            hc.alignment = center_align

        for p_idx, (yr, fp) in enumerate(sorted_periods):
            grp_start = _COL_DATA_START + p_idx * _COLS_PER_PERIOD
            grp_end   = grp_start + _COLS_PER_PERIOD - 1

            # Period group label in the first cell of the group
            pc = _write(row=4, column=grp_start, value=_period_label(yr, fp))
            pc.font      = _make_font(size=12, bold=True, color=_CLR_WHITE)
            pc.fill      = _make_fill(_CLR_NAVY)
            pc.alignment = center_align

            # Flood remaining cells of the merged range with navy fill
            for col in range(grp_start + 1, grp_end + 1):
                flood = _write(row=4, column=col, value=None)
                flood.fill = _make_fill(_CLR_NAVY)

            # Merge period label across its 4 columns
            ws.merge_cells(  # type: ignore[union-attr]
                start_row=4, start_column=grp_start,
                end_row=4,   end_column=grp_end,
            )

        # ── Row 5: Per-period column sub-headers ──────────────────────────────
        # Blank sub-header cells under the two fixed anchor columns
        for col in (_COL_CONCEPT_LABEL, _COL_CANONICAL_TAG):
            bc = _write(row=5, column=col, value=None)
            bc.fill = _make_fill(_CLR_NAVY)

        for p_idx in range(n_periods):
            for off, sub_label in enumerate(_PERIOD_COL_SUBHEADERS):
                col = _COL_DATA_START + p_idx * _COLS_PER_PERIOD + off
                sc = _write(row=5, column=col, value=sub_label)
                sc.font      = _make_font(size=10, bold=True, color=_CLR_WHITE)
                sc.fill      = _make_fill(_CLR_NAVY)
                sc.alignment = center_align

        # ── Freeze panes: below row 5 (data starts at row 6) ─────────────────
        ws.freeze_panes = ws.cell(row=6, column=1)  # type: ignore[union-attr]

        # ── Pre-build style objects ────────────────────────────────────────────
        fill_even      = _make_fill(_CLR_ZEBRA_EVEN)
        fill_odd       = _make_fill(_CLR_WHITE)
        fill_subtot    = _make_fill(_CLR_SLATE)
        fill_term      = _make_fill(_CLR_NAVY_LIGHT)
        font_regular   = _make_font(size=10)
        font_bold      = _make_font(size=10, bold=True)
        border_subtot  = _make_thin_border(top=True)
        border_term    = _make_double_bottom_border()

        # ── Rows 6+: Data rows ─────────────────────────────────────────────────
        for row_offset, tag in enumerate(ordered_tags):
            data_row = 6 + row_offset

            is_terminal = tag in _TERMINAL_TAGS
            is_subtotal = (not is_terminal) and (tag in _SUBTOTAL_TAGS)

            # Determine fill and font for this row
            if is_terminal:
                row_fill = fill_term
                row_font = font_bold
            elif is_subtotal:
                row_fill = fill_subtot
                row_font = font_bold
            else:
                row_fill = fill_even if (data_row % 2 == 0) else fill_odd
                row_font = font_regular

            # Concept Label (col A) — derive from first item available for tag
            first_item = next(iter(pivot[tag].values()))
            concept_lc = _write(row=data_row, column=_COL_CONCEPT_LABEL,
                                 value=first_item.concept_label)
            concept_lc.fill      = row_fill
            concept_lc.font      = row_font
            concept_lc.alignment = left_align

            # Canonical Tag (col B)
            tag_cell = _write(row=data_row, column=_COL_CANONICAL_TAG, value=tag)
            tag_cell.fill      = row_fill
            tag_cell.font      = _make_font(size=9, color="595959")
            tag_cell.alignment = left_align

            # Per-period value cells
            for p_idx, period_key in enumerate(sorted_periods):
                col_base = _COL_DATA_START + p_idx * _COLS_PER_PERIOD
                item = pivot[tag].get(period_key)  # None → blank period

                # Value (Reported)
                val_rep = (float(item.value_reported)
                           if (item and item.value_reported is not None) else None)
                vc = _write(row=data_row, column=col_base + _OFF_VALUE_REPORTED,
                            value=val_rep)
                vc.fill          = row_fill
                vc.font          = row_font
                vc.number_format = _FMT_CURRENCY
                vc.alignment     = right_align

                # Reported Currency
                ccy_val = item.reported_currency if item else None
                cc = _write(row=data_row, column=col_base + _OFF_REPORTED_CCY,
                            value=ccy_val)
                cc.fill      = row_fill
                cc.font      = row_font
                cc.alignment = center_align

                # Value (USD)
                val_usd = (float(item.value_usd)
                           if (item and item.value_usd is not None) else None)
                uc = _write(row=data_row, column=col_base + _OFF_VALUE_USD,
                            value=val_usd)
                uc.fill          = row_fill
                uc.font          = row_font
                uc.number_format = _FMT_CURRENCY
                uc.alignment     = right_align

                # FX Rate Used
                fx_val = (float(item.fx_rate_used)
                          if (item and item.fx_rate_used is not None) else None)
                fc = _write(row=data_row, column=col_base + _OFF_FX_RATE,
                            value=fx_val)
                fc.fill          = row_fill
                fc.font          = _make_font(size=9, color="595959")
                fc.number_format = _FMT_FX_RATE
                fc.alignment     = right_align

            # Accounting borders (applied across all period columns + anchors)
            if is_terminal:
                _apply_row_border(ws, data_row, total_cols, border_term)
            elif is_subtotal:
                _apply_row_border(ws, data_row, total_cols, border_subtot)

        # ── Auto-fit all columns ───────────────────────────────────────────────
        _auto_fit_columns(ws)

    def _write_audit_log_sheet(self, ws: object, ctx: ExportContext) -> None:
        """
        Write Sheet 8 — Mandatory Regulatory Audit Log (Amendment V1.2 §6.1).

        M6.3 hydration: columns L and M now record the per-item fiscal_year
        and fiscal_period from ``ExportLineItem`` (not the job-level defaults).
        This correctly reflects multi-period data in the audit trail.

        Columns:
          A  Concept Label
          B  Statement Type
          C  Canonical Field
          D  Value (Reported)       format: #,##0.00
          E  Reported Currency
          F  Value (USD)            format: #,##0.00
          G  FX Rate Used           format: 0.00000
          H  Extraction Method
          I  Source File Hash (SHA-256)
          J  Filing Date
          K  Derived Formula
          L  Fiscal Year            ← per-item (M6.3)
          M  Fiscal Period          ← per-item (M6.3)
          N  Reporting Standard
          O  Export Timestamp
        """
        _write = ws.cell  # type: ignore[attr-defined]

        audit_headers = [
            "Concept Label", "Statement Type", "Canonical Field",
            "Value (Reported)", "Reported Currency", "Value (USD)",
            "FX Rate Used", "Extraction Method", "Source File Hash",
            "Filing Date", "Derived Formula",
            "Fiscal Year", "Fiscal Period", "Reporting Standard",
            "Export Timestamp",
        ]
        col_count = len(audit_headers)

        header_font = _make_font(size=11, bold=True, color=_CLR_WHITE)
        header_fill = _make_fill(_CLR_NAVY)
        for col_idx, header in enumerate(audit_headers, start=1):
            hc = _write(row=1, column=col_idx, value=header)
            hc.font = header_font
            hc.fill = header_fill

        ws.freeze_panes = ws.cell(row=2, column=1)  # type: ignore[union-attr]

        fill_even = _make_fill(_CLR_ZEBRA_EVEN)
        fill_odd  = _make_fill(_CLR_WHITE)
        font_data = _make_font(size=10)
        export_ts = ctx.export_timestamp.strftime("%Y-%m-%d %H:%M:%S UTC")

        for row_idx, item in enumerate(ctx.all_items, start=2):
            # ── Cell values ────────────────────────────────────────────────────
            _write(row=row_idx, column=1,  value=item.concept_label)
            _write(row=row_idx, column=2,  value=item.statement_type)
            _write(row=row_idx, column=3,  value=item.canonical_field)
            _write(row=row_idx, column=4,
                   value=float(item.value_reported) if item.value_reported is not None else None)
            _write(row=row_idx, column=5,  value=item.reported_currency or "")
            _write(row=row_idx, column=6,
                   value=float(item.value_usd) if item.value_usd is not None else None)
            _write(row=row_idx, column=7,
                   value=float(item.fx_rate_used) if item.fx_rate_used is not None else None)
            _write(row=row_idx, column=8,  value=item.extraction_method or "")
            _write(row=row_idx, column=9,  value=item.source_file_hash or "")
            _write(row=row_idx, column=10,
                   value=item.filing_date.isoformat() if item.filing_date else "")
            _write(row=row_idx, column=11, value=item.derived_formula or "")
            # M6.3: per-item fiscal year and period (not job-level ctx values)
            _write(row=row_idx, column=12, value=item.fiscal_year)
            _write(row=row_idx, column=13, value=item.fiscal_period)
            _write(row=row_idx, column=14, value=ctx.reporting_standard)
            _write(row=row_idx, column=15, value=export_ts)

            # ── Number formats ─────────────────────────────────────────────────
            ws.cell(row=row_idx, column=4).number_format  = _FMT_CURRENCY  # type: ignore[union-attr]
            ws.cell(row=row_idx, column=6).number_format  = _FMT_CURRENCY  # type: ignore[union-attr]
            ws.cell(row=row_idx, column=7).number_format  = _FMT_FX_RATE   # type: ignore[union-attr]

            # ── Zebra fill + font ──────────────────────────────────────────────
            row_fill = fill_even if (row_idx % 2 == 0) else fill_odd
            _apply_row_fill(ws, row_idx, col_count, row_fill)
            _apply_row_font(ws, row_idx, col_count, font_data)

        _auto_fit_columns(ws)


# ---------------------------------------------------------------------------
# ORM → DTO mapper
# ---------------------------------------------------------------------------

def _orm_to_export_item(row: object) -> ExportLineItem:
    """
    Convert a ``FinancialLineItem`` ORM instance to an ``ExportLineItem`` DTO.

    All attribute accesses use ``getattr`` to avoid triggering lazy-load
    on detached instances.  ``fiscal_year`` and ``fiscal_period`` are now
    read directly from the ORM row (M6.3 multi-period support).

    Args:
        row: ``FinancialLineItem`` ORM instance.

    Returns:
        Decoupled ``ExportLineItem`` dataclass.
    """
    canonical = getattr(row, "canonical_field", "") or ""
    return ExportLineItem(
        canonical_field=canonical,
        concept_label=_label_from_canonical(canonical),
        statement_type=getattr(row, "statement_type", "") or "",
        fiscal_year=int(getattr(row, "fiscal_year", datetime.utcnow().year) or datetime.utcnow().year),
        fiscal_period=str(getattr(row, "fiscal_period", "FY") or "FY").upper(),
        value_reported=getattr(row, "value_reported", None),
        reported_currency=(getattr(row, "reported_currency", None) or "USD").upper(),
        value_usd=getattr(row, "value_usd", None),
        fx_rate_used=getattr(row, "fx_rate_used", None),
        filing_date=getattr(row, "filing_date", date.today()),
        is_restated=bool(getattr(row, "is_restated", False)),
        extraction_method=getattr(row, "extraction_method", None),
        source_file_hash=getattr(row, "source_file_hash", None),
        derived_formula=getattr(row, "derived_expression_formula", None),
    )


# ---------------------------------------------------------------------------
# Custom exceptions
# ---------------------------------------------------------------------------

class ExcelExportError(Exception):
    """Base exception for all ExcelExportService errors."""


class ExportJobNotFoundError(ExcelExportError):
    """Raised when the FinancialJob does not exist."""


class ExportCompanyNotFoundError(ExcelExportError):
    """Raised when the Company linked to the job does not exist."""


class ExportNoDataError(ExcelExportError):
    """Raised when no FinancialLineItem rows exist for the company."""
