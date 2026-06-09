"""
Sheet builder: Financial Ratios (Sheet 5).

Computes seven key financial ratios from the Income Statement and Balance
Sheet items already loaded into ``ExportContext``.  Each ratio is presented
as a single row with one value column per chronological fiscal period.

Ratios computed
───────────────
  Profitability
    1. Gross Margin %          = GrossProfit / Revenue × 100
    2. Operating Margin %      = OperatingIncomeLoss / Revenue × 100
    3. Net Margin %            = NetIncomeLoss / Revenue × 100
    4. Return on Equity (ROE)% = NetIncomeLoss / StockholdersEquity × 100

  Liquidity
    5. Current Ratio           = AssetsCurrent / LiabilitiesCurrent

  Leverage
    6. Debt-to-Equity          = (LiabilitiesCurrent + LiabilitiesNoncurrent)
                                  / StockholdersEquity

  Efficiency
    7. Asset Turnover          = Revenue / TotalAssets

Layout
──────
  Row 1:   Title banner (navy fill, white bold)
  Row 2:   Methodology note (slate fill)
  Row 3:   Blank separator
  Row 4:   Column headers (navy fill, white bold)
  Row 5+:  Category separator rows (slate) interspersed with ratio rows (zebra)

  Fixed columns:
    Col A  — Ratio Name
    Col B  — Category
    Col C  — Formula (abbreviated)
  Dynamic columns (Col D onward):
    One column per chronological (fiscal_year, fiscal_period) pair.

Number format
─────────────
  Percentage ratios: "0.00%" (stored as decimal fraction, e.g. 0.4212)
  Multiple ratios:   "0.00"

Missing values (denominator zero or numerator absent) are rendered as "—".

Milestone: B6 — Advanced Excel Sheet Completion.
"""

from __future__ import annotations

from decimal import Decimal, DivisionByZero, InvalidOperation
from typing import Final

# ---------------------------------------------------------------------------
# Canonical tag alias sets — maps each concept to the set of XBRL tags that
# represent it across US-GAAP, IFRS, and IND-AS reporting standards.
# ---------------------------------------------------------------------------

_REVENUE_TAGS: Final[frozenset[str]] = frozenset({
    "us-gaap:Revenues",
    "us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax",
    "ifrs-full:Revenue",
    "ifrs-full:RevenueFromContractsWithCustomers",
    "ind-as:Revenue",
    "ind-as:RevenueFromOperations",
    "ind-as:TotalIncome",
})

_GROSS_PROFIT_TAGS: Final[frozenset[str]] = frozenset({
    "us-gaap:GrossProfit",
    "ifrs-full:GrossProfit",
    "ind-as:GrossProfit",
})

_OPERATING_INCOME_TAGS: Final[frozenset[str]] = frozenset({
    "us-gaap:OperatingIncomeLoss",
    "ifrs-full:ProfitLossFromOperatingActivities",
    "ind-as:ProfitBeforeExceptionalItemsAndTax",
    "ind-as:ProfitFromOperations",
})

_NET_INCOME_TAGS: Final[frozenset[str]] = frozenset({
    "us-gaap:NetIncomeLoss",
    "ifrs-full:ProfitLoss",
    "ind-as:ProfitLoss",
    "ind-as:ProfitForYear",
})

_CURRENT_ASSETS_TAGS: Final[frozenset[str]] = frozenset({
    "us-gaap:AssetsCurrent",
    "ifrs-full:CurrentAssets",
    "ind-as:CurrentAssets",
})

_CURRENT_LIABILITIES_TAGS: Final[frozenset[str]] = frozenset({
    "us-gaap:LiabilitiesCurrent",
    "ifrs-full:CurrentLiabilities",
    "ind-as:CurrentLiabilities",
})

_TOTAL_ASSETS_TAGS: Final[frozenset[str]] = frozenset({
    "us-gaap:Assets",
    "ifrs-full:Assets",
    "ind-as:Assets",
    "ind-as:TotalAssets",
})

_NONCURRENT_LIABILITIES_TAGS: Final[frozenset[str]] = frozenset({
    "us-gaap:LiabilitiesNoncurrent",
    "ifrs-full:NoncurrentLiabilities",
    "ind-as:NoncurrentLiabilities",
})

_EQUITY_TAGS: Final[frozenset[str]] = frozenset({
    "us-gaap:StockholdersEquity",
    "us-gaap:StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest",
    "ifrs-full:Equity",
    "ind-as:Equity",
    "ind-as:TotalEquity",
})

# ---------------------------------------------------------------------------
# Ratio definitions — drives both computation and rendering order.
# Entries with type == "__SEPARATOR__" render as category divider rows.
# ---------------------------------------------------------------------------

_SEPARATOR: Final[str] = "__SEPARATOR__"

_RATIO_DEFS: Final[list[dict]] = [
    # ── Profitability ────────────────────────────────────────────────────────
    {"type": _SEPARATOR, "label": "Profitability"},
    {
        "label":    "Gross Margin",
        "category": "Profitability",
        "formula":  "Gross Profit / Revenue",
        "pct":      True,
        "compute":  "gross_margin",
    },
    {
        "label":    "Operating Margin",
        "category": "Profitability",
        "formula":  "Operating Income / Revenue",
        "pct":      True,
        "compute":  "operating_margin",
    },
    {
        "label":    "Net Margin",
        "category": "Profitability",
        "formula":  "Net Income / Revenue",
        "pct":      True,
        "compute":  "net_margin",
    },
    {
        "label":    "Return on Equity (ROE)",
        "category": "Profitability",
        "formula":  "Net Income / Total Equity",
        "pct":      True,
        "compute":  "roe",
    },
    # ── Liquidity ────────────────────────────────────────────────────────────
    {"type": _SEPARATOR, "label": "Liquidity"},
    {
        "label":    "Current Ratio",
        "category": "Liquidity",
        "formula":  "Current Assets / Current Liabilities",
        "pct":      False,
        "compute":  "current_ratio",
    },
    # ── Leverage ─────────────────────────────────────────────────────────────
    {"type": _SEPARATOR, "label": "Leverage"},
    {
        "label":    "Debt-to-Equity",
        "category": "Leverage",
        "formula":  "(Current + Non-current Liabilities) / Equity",
        "pct":      False,
        "compute":  "debt_to_equity",
    },
    # ── Efficiency ───────────────────────────────────────────────────────────
    {"type": _SEPARATOR, "label": "Efficiency"},
    {
        "label":    "Asset Turnover",
        "category": "Efficiency",
        "formula":  "Revenue / Total Assets",
        "pct":      False,
        "compute":  "asset_turnover",
    },
]

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _safe_div(
    numerator: Decimal | None,
    denominator: Decimal | None,
) -> Decimal | None:
    """Return ``numerator / denominator``, or ``None`` on missing / zero denominator."""
    if numerator is None or denominator is None:
        return None
    try:
        if denominator == Decimal(0):
            return None
        return numerator / denominator
    except (DivisionByZero, InvalidOperation):
        return None


def _lookup(
    pivot: dict[tuple[int, str], dict[str, Decimal]],
    period: tuple[int, str],
    tags: frozenset[str],
) -> Decimal | None:
    """
    Return the first non-None USD value found for any of *tags* in *period*.

    The pivot maps ``(fiscal_year, fiscal_period) → {canonical_field: value_usd}``.
    """
    period_data = pivot.get(period, {})
    for tag in tags:
        val = period_data.get(tag)
        if val is not None:
            return val
    return None


def _compute_ratios(
    pivot: dict[tuple[int, str], dict[str, Decimal]],
    period: tuple[int, str],
) -> dict[str, Decimal | None]:
    """Compute all seven ratios for a single *period* from *pivot*."""
    revenue     = _lookup(pivot, period, _REVENUE_TAGS)
    gross_p     = _lookup(pivot, period, _GROSS_PROFIT_TAGS)
    op_income   = _lookup(pivot, period, _OPERATING_INCOME_TAGS)
    net_income  = _lookup(pivot, period, _NET_INCOME_TAGS)
    curr_assets = _lookup(pivot, period, _CURRENT_ASSETS_TAGS)
    curr_liab   = _lookup(pivot, period, _CURRENT_LIABILITIES_TAGS)
    total_ass   = _lookup(pivot, period, _TOTAL_ASSETS_TAGS)
    nc_liab     = _lookup(pivot, period, _NONCURRENT_LIABILITIES_TAGS)
    equity      = _lookup(pivot, period, _EQUITY_TAGS)

    # Total liabilities = current + non-current (fall back to current only)
    total_liab: Decimal | None = None
    if curr_liab is not None and nc_liab is not None:
        total_liab = curr_liab + nc_liab
    elif curr_liab is not None:
        total_liab = curr_liab
    elif nc_liab is not None:
        total_liab = nc_liab

    return {
        "gross_margin":     _safe_div(gross_p, revenue),
        "operating_margin": _safe_div(op_income, revenue),
        "net_margin":       _safe_div(net_income, revenue),
        "roe":              _safe_div(net_income, equity),
        "current_ratio":    _safe_div(curr_assets, curr_liab),
        "debt_to_equity":   _safe_div(total_liab, equity),
        "asset_turnover":   _safe_div(revenue, total_ass),
    }


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def write_ratios_sheet(ws: object, ctx: object) -> None:  # type: ignore[type-arg]
    """
    Populate the Financial Ratios worksheet *ws* from *ctx*.

    Called synchronously from ``ExcelExportService._build_workbook`` after the
    four statement sheets have been written.  The openpyxl ``Worksheet`` is
    already created by the caller; this function fills it.

    Parameters
    ----------
    ws:
        An openpyxl ``Worksheet`` object.
    ctx:
        Populated ``ExportContext`` carrying ``is_items``, ``bs_items``, and
        the ``periods`` property.
    """
    # Deferred imports — keeps module importable without openpyxl installed.
    from openpyxl.styles import Alignment  # type: ignore[import]
    from openpyxl.utils import get_column_letter  # type: ignore[import]

    # Import styling helpers from the parent module at call-time to avoid
    # circular imports and mirror the deferred-import pattern used throughout.
    from services.export.excel_generator import (  # type: ignore[import]
        _CLR_NAVY,
        _CLR_SLATE,
        _CLR_ZEBRA_EVEN,
        _CLR_WHITE,
        _CLR_BLACK,
        _make_fill,
        _make_font,
        _apply_row_fill,
        _apply_row_font,
        _auto_fit_columns,
        _period_label,
        ExportContext,
    )

    ctx_typed: ExportContext = ctx  # type: ignore[assignment]

    # ── Build period-keyed USD value pivot ────────────────────────────────────
    # pivot: (fiscal_year, fiscal_period) → {canonical_field: value_usd}
    # Uses value_usd for cross-currency comparability.  Items without value_usd
    # are skipped.  Later items overwrite earlier ones (most-recent-filing wins).
    pivot: dict[tuple[int, str], dict[str, Decimal]] = {}
    for item in ctx_typed.is_items + ctx_typed.bs_items:
        if item.value_usd is None:
            continue
        bucket = pivot.setdefault(item.period_key, {})
        bucket[item.canonical_field] = item.value_usd

    periods = ctx_typed.periods  # list[tuple[int, str]]

    # ── Column indices ────────────────────────────────────────────────────────
    _COL_NAME     = 1
    _COL_CATEGORY = 2
    _COL_FORMULA  = 3
    _COL_DATA     = 4
    total_cols    = (_COL_DATA + len(periods) - 1) if periods else _COL_DATA

    # ── Fill / font / alignment objects ──────────────────────────────────────
    fill_navy    = _make_fill(_CLR_NAVY)
    fill_slate   = _make_fill(_CLR_SLATE)
    fill_even    = _make_fill(_CLR_ZEBRA_EVEN)
    fill_white   = _make_fill(_CLR_WHITE)

    font_header    = _make_font(size=11, bold=True,  color="FFFFFF")
    font_sub       = _make_font(size=10, bold=False, color=_CLR_BLACK)
    font_label     = _make_font(size=10, bold=True,  color=_CLR_BLACK)
    font_separator = _make_font(size=10, bold=True,  color=_CLR_BLACK)
    font_data      = _make_font(size=10, bold=False, color=_CLR_BLACK)
    font_note      = _make_font(size=9,  bold=False, color="595959")

    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")
    right_align  = Alignment(horizontal="right",  vertical="center")

    # ── Row 1: Title banner ───────────────────────────────────────────────────
    _apply_row_fill(ws, 1, total_cols, fill_navy)
    _apply_row_font(ws, 1, total_cols, font_header)
    title_cell = ws.cell(  # type: ignore[union-attr]
        row=1, column=_COL_NAME,
        value=(
            f"Financial Ratios — {ctx_typed.company_name}"
            f" ({ctx_typed.company_ticker})"
            f"  ·  {ctx_typed.period_range_label}"
        ),
    )
    title_cell.font      = font_header
    title_cell.fill      = fill_navy
    title_cell.alignment = left_align

    # ── Row 2: Methodology note ───────────────────────────────────────────────
    _apply_row_fill(ws, 2, total_cols, fill_slate)
    note_cell = ws.cell(  # type: ignore[union-attr]
        row=2, column=_COL_NAME,
        value=(
            "All monetary inputs translated to USD at rates stored during "
            "extraction.  '—' denotes insufficient source data for the period."
        ),
    )
    note_cell.font      = font_note
    note_cell.fill      = fill_slate
    note_cell.alignment = left_align

    # ── Row 3: Blank separator ────────────────────────────────────────────────
    _apply_row_fill(ws, 3, total_cols, fill_white)

    # ── Row 4: Column headers ─────────────────────────────────────────────────
    header_row = 4
    _apply_row_fill(ws, header_row, total_cols, fill_navy)
    _apply_row_font(ws, header_row, total_cols, font_header)

    ws.cell(row=header_row, column=_COL_NAME,     value="Ratio").alignment     = center_align  # type: ignore[union-attr]
    ws.cell(row=header_row, column=_COL_CATEGORY, value="Category").alignment  = center_align  # type: ignore[union-attr]
    ws.cell(row=header_row, column=_COL_FORMULA,  value="Formula").alignment   = center_align  # type: ignore[union-attr]

    for p_idx, period in enumerate(periods):
        col  = _COL_DATA + p_idx
        cell = ws.cell(row=header_row, column=col, value=_period_label(*period))  # type: ignore[union-attr]
        cell.font      = font_header
        cell.fill      = fill_navy
        cell.alignment = center_align

    # ── Pre-compute ratios for all periods ────────────────────────────────────
    period_ratios: dict[tuple[int, str], dict[str, Decimal | None]] = {
        p: _compute_ratios(pivot, p) for p in periods
    }

    # ── Data rows (Row 5 onward) ──────────────────────────────────────────────
    row_idx  = 5
    data_row = 0   # 0-based counter for zebra alternation

    for defn in _RATIO_DEFS:
        if defn.get("type") == _SEPARATOR:
            # Category separator row
            _apply_row_fill(ws, row_idx, total_cols, fill_slate)
            _apply_row_font(ws, row_idx, total_cols, font_separator)
            sep_cell = ws.cell(row=row_idx, column=_COL_NAME, value=defn["label"])  # type: ignore[union-attr]
            sep_cell.font      = font_separator
            sep_cell.fill      = fill_slate
            sep_cell.alignment = left_align
            row_idx  += 1
            data_row  = 0
            continue

        # Ratio value row
        row_fill = fill_even if (data_row % 2 == 0) else fill_white
        _apply_row_fill(ws, row_idx, total_cols, row_fill)

        name_cell = ws.cell(row=row_idx, column=_COL_NAME, value=defn["label"])  # type: ignore[union-attr]
        name_cell.font      = font_label
        name_cell.fill      = row_fill
        name_cell.alignment = left_align

        cat_cell = ws.cell(row=row_idx, column=_COL_CATEGORY, value=defn["category"])  # type: ignore[union-attr]
        cat_cell.font      = font_sub
        cat_cell.fill      = row_fill
        cat_cell.alignment = left_align

        fml_cell = ws.cell(row=row_idx, column=_COL_FORMULA, value=defn["formula"])  # type: ignore[union-attr]
        fml_cell.font      = font_sub
        fml_cell.fill      = row_fill
        fml_cell.alignment = left_align

        is_pct      = defn["pct"]
        compute_key: str = defn["compute"]

        for p_idx, period in enumerate(periods):
            col   = _COL_DATA + p_idx
            value = period_ratios[period].get(compute_key)

            if value is None:
                cell = ws.cell(row=row_idx, column=col, value="—")  # type: ignore[union-attr]
                cell.font      = font_data
                cell.fill      = row_fill
                cell.alignment = center_align
            else:
                cell = ws.cell(row=row_idx, column=col)  # type: ignore[union-attr]
                cell.value         = float(value)
                cell.number_format = "0.00%" if is_pct else "0.00"
                cell.font          = font_data
                cell.fill          = row_fill
                cell.alignment     = right_align

        row_idx  += 1
        data_row += 1

    # ── No-data fallback ──────────────────────────────────────────────────────
    if not periods:
        empty_cell = ws.cell(  # type: ignore[union-attr]
            row=row_idx, column=_COL_NAME,
            value="No financial data available for this export.",
        )
        empty_cell.font      = _make_font(size=10, color="808080")
        empty_cell.alignment = left_align

    # ── Column widths ─────────────────────────────────────────────────────────
    _auto_fit_columns(ws)
    ws.column_dimensions[get_column_letter(_COL_NAME)].width     = 36  # type: ignore[union-attr]
    ws.column_dimensions[get_column_letter(_COL_CATEGORY)].width = 16  # type: ignore[union-attr]
    ws.column_dimensions[get_column_letter(_COL_FORMULA)].width  = 40  # type: ignore[union-attr]
