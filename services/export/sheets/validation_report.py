"""
Sheet builder: Validation Report — QA Pipeline Summary (Sheet 9).

Surfaces the full quality-assurance audit record for the export's originating
ingestion job.  Data is sourced from the ``validation_data`` field added to
``ExportContext`` (populated from the ``validation_results`` table by
``_load_export_context``).

If no validation record is available (e.g. the pipeline ran before M4 was
deployed), a graceful "No validation record found" notice is written instead.

Sections rendered
──────────────────
  1. Validation Summary     — confidence score, item count, exportable flag,
                              critical/warning counts, accession number, period
  2. Confidence Score       — prominent score cell (colour-coded: green ≥80,
                              amber 50–79, red <50) + deductions table
  3. Anomaly Findings       — full itemised ledger of findings, sorted:
                              CRITICAL → WARNING → INFO; each row shows
                              rule_id, severity, message, expected, actual, delta
  4. Summary Text           — raw pipeline summary note (if present)

Severity colour coding
──────────────────────
  CRITICAL  — red background    (FFE4E4)
  WARNING   — amber background  (FFF3CD)
  INFO      — slate background  (D9E1F2)

Milestone: B6 — Advanced Excel Sheet Completion.
"""

from __future__ import annotations

from typing import Any, Final

_SEV_ORDER: Final[dict[str, int]] = {"CRITICAL": 0, "WARNING": 1, "INFO": 2}
_SEV_FILL: Final[dict[str, str]] = {
    "CRITICAL": "FFE4E4",
    "WARNING":  "FFF3CD",
    "INFO":     "D9E1F2",
}


def _score_fill(score: int) -> str:
    """Background hex for a confidence score cell."""
    if score >= 80:
        return "D4EDDA"   # green
    if score >= 50:
        return "FFF3CD"   # amber
    return "FFE4E4"        # red


def _score_font_color(score: int) -> str:
    if score >= 80:
        return "155724"
    if score >= 50:
        return "856404"
    return "721C24"


def write_validation_report_sheet(ws: object, ctx: object) -> None:  # type: ignore[type-arg]
    """
    Populate the Validation Report worksheet from *ctx*.

    Called synchronously from ``ExcelExportService._build_workbook``.
    ``ctx.validation_data`` may be ``None`` if no validation record exists for
    this job — the sheet is written with a graceful notice in that case.
    """
    from openpyxl.styles import Alignment  # type: ignore[import]
    from openpyxl.utils import get_column_letter  # type: ignore[import]

    from services.export.excel_generator import (  # type: ignore[import]
        _CLR_NAVY,
        _CLR_NAVY_LIGHT,
        _CLR_SLATE,
        _CLR_ZEBRA_EVEN,
        _CLR_WHITE,
        _CLR_BLACK,
        _make_fill,
        _make_font,
        _apply_row_fill,
        _apply_row_font,
        ExportContext,
    )

    ctx_typed: ExportContext = ctx  # type: ignore[assignment]
    vdata: dict[str, Any] | None = getattr(ctx_typed, "validation_data", None)

    TOTAL_COLS = 6

    fill_navy       = _make_fill(_CLR_NAVY)
    fill_navy_light = _make_fill(_CLR_NAVY_LIGHT)
    fill_slate      = _make_fill(_CLR_SLATE)
    fill_even       = _make_fill(_CLR_ZEBRA_EVEN)
    fill_white      = _make_fill(_CLR_WHITE)

    font_banner = _make_font(size=13, bold=True,  color="FFFFFF")
    font_sec    = _make_font(size=11, bold=True,  color=_CLR_BLACK)
    font_label  = _make_font(size=10, bold=True,  color=_CLR_BLACK)
    font_value  = _make_font(size=10, bold=False, color=_CLR_BLACK)
    font_note   = _make_font(size=9,  bold=False, color="595959")
    font_mono   = _make_font(size=9,  bold=False, color=_CLR_BLACK)

    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    right_align  = Alignment(horizontal="right",  vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    row = 1

    def _blank() -> None:
        nonlocal row
        _apply_row_fill(ws, row, TOTAL_COLS, fill_white)
        row += 1

    def _section(title: str) -> None:
        nonlocal row
        _apply_row_fill(ws, row, TOTAL_COLS, fill_slate)
        c = ws.cell(row=row, column=1, value=title)  # type: ignore[union-attr]
        c.font = font_sec; c.fill = fill_slate; c.alignment = left_align
        row += 1

    def _kv(label: str, value: str, even: bool) -> None:
        nonlocal row
        rf = fill_even if even else fill_white
        _apply_row_fill(ws, row, TOTAL_COLS, rf)
        lc = ws.cell(row=row, column=1, value=label)  # type: ignore[union-attr]
        lc.font = font_label; lc.fill = rf; lc.alignment = left_align
        vc = ws.cell(row=row, column=2, value=value)  # type: ignore[union-attr]
        vc.font = font_value; vc.fill = rf; vc.alignment = left_align
        row += 1

    # ── Row 1: Title banner ───────────────────────────────────────────────────
    _apply_row_fill(ws, row, TOTAL_COLS, fill_navy)
    _apply_row_font(ws, row, TOTAL_COLS, font_banner)
    tc = ws.cell(row=row, column=1,  # type: ignore[union-attr]
        value=(
            f"Validation Report — {ctx_typed.company_name}"
            f" ({ctx_typed.company_ticker})  ·  {ctx_typed.period_range_label}"
        ),
    )
    tc.font = font_banner; tc.fill = fill_navy; tc.alignment = left_align
    row += 1

    # ── Row 2: Note ───────────────────────────────────────────────────────────
    _apply_row_fill(ws, row, TOTAL_COLS, fill_slate)
    nc = ws.cell(row=row, column=1,  # type: ignore[union-attr]
        value=(
            "Quality-assurance results from the dual-dimension validation engine "
            "(Amendment V1.2 §5).  Confidence score deducted 25 pts per CRITICAL "
            "finding, 5 pts per WARNING."
        ),
    )
    nc.font = font_note; nc.fill = fill_slate; nc.alignment = left_align
    row += 1
    _blank()

    # ── No validation data guard ──────────────────────────────────────────────
    if vdata is None:
        _apply_row_fill(ws, row, TOTAL_COLS, fill_even)
        c = ws.cell(row=row, column=1,  # type: ignore[union-attr]
            value=(
                "No validation record found for this job.  "
                "Run the extraction pipeline to generate QA data."
            ),
        )
        c.font = _make_font(size=10, color="808080")
        c.alignment = left_align
        ws.column_dimensions[get_column_letter(1)].width = 72  # type: ignore[union-attr]
        return

    # ── Section 1: Validation Summary ────────────────────────────────────────
    _section("1. Validation Summary")
    score     = int(vdata.get("confidence_score", 0))
    exportable = vdata.get("is_exportable", False)
    summary_rows: list[tuple[str, str]] = [
        ("Accession Number",    str(vdata.get("accession_number", "—"))),
        ("Fiscal Year",         str(vdata.get("fiscal_year") or "—")),
        ("Fiscal Period",       str(vdata.get("fiscal_period") or "—")),
        ("Items Validated",     str(vdata.get("items_validated", 0))),
        ("Critical Findings",   str(vdata.get("critical_count", 0))),
        ("Warning Findings",    str(vdata.get("warning_count", 0))),
        ("Export Gate",         "✓ READY FOR EXPORT" if exportable else "✗ EXPORT BLOCKED"),
        ("Validation ID",       str(vdata.get("id", "—"))),
        ("Validated At",        str(vdata.get("created_at", "—"))),
    ]
    for i, (lbl, val) in enumerate(summary_rows):
        _kv(lbl, val, i % 2 == 0)
    _blank()

    # ── Section 2: Confidence Score ───────────────────────────────────────────
    _section("2. Confidence Score")
    score_fill_hex  = _score_fill(score)
    score_font_hex  = _score_font_color(score)

    # Big score row
    _apply_row_fill(ws, row, TOTAL_COLS, _make_fill(score_fill_hex))
    sc = ws.cell(row=row, column=1, value="Confidence Score")  # type: ignore[union-attr]
    sc.font      = _make_font(size=11, bold=True, color=score_font_hex)
    sc.fill      = _make_fill(score_fill_hex)
    sc.alignment = left_align
    sv = ws.cell(row=row, column=2, value=f"{score} / 100")  # type: ignore[union-attr]
    sv.font      = _make_font(size=14, bold=True, color=score_font_hex)
    sv.fill      = _make_fill(score_fill_hex)
    sv.alignment = center_align
    row += 1
    _blank()

    # Deductions table
    deductions: list[dict[str, Any]] = vdata.get("deductions") or []
    if deductions:
        _apply_row_fill(ws, row, TOTAL_COLS, fill_navy_light)
        for col, hdr in [(1, "Rule ID"), (2, "Points Deducted"), (3, "Reason")]:
            c = ws.cell(row=row, column=col, value=hdr)  # type: ignore[union-attr]
            c.font = font_label; c.fill = fill_navy_light; c.alignment = center_align
        row += 1
        for i, ded in enumerate(deductions):
            rf = fill_even if (i % 2 == 0) else fill_white
            _apply_row_fill(ws, row, TOTAL_COLS, rf)
            for col, val in [
                (1, str(ded.get("rule_id", ""))),
                (2, str(ded.get("points", ""))),
                (3, str(ded.get("reason", ""))),
            ]:
                c = ws.cell(row=row, column=col, value=val)  # type: ignore[union-attr]
                c.font = font_mono; c.fill = rf
                c.alignment = right_align if col == 2 else left_align
            row += 1
    _blank()

    # ── Section 3: Anomaly Findings ───────────────────────────────────────────
    _section("3. Anomaly Findings Ledger")
    findings: list[dict[str, Any]] = vdata.get("findings") or []
    if not findings:
        _kv("Finding count", "0 — no anomalies detected", True)
    else:
        sorted_findings = sorted(
            findings, key=lambda f: _SEV_ORDER.get(str(f.get("severity", "INFO")).upper(), 99)
        )

        # Column header row
        _apply_row_fill(ws, row, TOTAL_COLS, fill_navy_light)
        for col, hdr in [
            (1, "Rule ID"), (2, "Severity"), (3, "Message"),
            (4, "Expected"), (5, "Actual"), (6, "Delta"),
        ]:
            c = ws.cell(row=row, column=col, value=hdr)  # type: ignore[union-attr]
            c.font = font_label; c.fill = fill_navy_light; c.alignment = center_align
        row += 1

        for finding in sorted_findings:
            sev     = str(finding.get("severity", "INFO")).upper()
            sev_hex = _SEV_FILL.get(sev, _CLR_ZEBRA_EVEN)
            rf      = _make_fill(sev_hex)
            _apply_row_fill(ws, row, TOTAL_COLS, rf)

            for col, val in [
                (1, str(finding.get("rule_id", ""))),
                (2, sev),
                (3, str(finding.get("message", ""))),
                (4, _fmt_num(finding.get("expected"))),
                (5, _fmt_num(finding.get("actual"))),
                (6, _fmt_num(finding.get("delta"))),
            ]:
                c = ws.cell(row=row, column=col, value=val)  # type: ignore[union-attr]
                c.font      = font_mono
                c.fill      = rf
                c.alignment = right_align if col >= 4 else left_align
            row += 1
    _blank()

    # ── Section 4: Summary Text ───────────────────────────────────────────────
    summary_text: str | None = vdata.get("summary_text")
    if summary_text:
        _section("4. Pipeline Summary Note")
        _apply_row_fill(ws, row, TOTAL_COLS, fill_even)
        sc2 = ws.cell(row=row, column=1, value=summary_text)  # type: ignore[union-attr]
        sc2.font      = font_value
        sc2.fill      = fill_even
        sc2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = max(60, len(summary_text) // 6)  # type: ignore[union-attr]
        row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = {1: 20, 2: 18, 3: 56, 4: 16, 5: 16, 6: 16}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width  # type: ignore[union-attr]


def _fmt_num(v: Any) -> str:
    """Format a numeric finding field as a compact string."""
    if v is None:
        return "—"
    try:
        f = float(v)
        if f == int(f):
            return f"{int(f):,}"
        return f"{f:,.4f}"
    except (TypeError, ValueError):
        return str(v)
