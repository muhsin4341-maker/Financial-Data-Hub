"""
Sheet builder: Strategic Initiatives — Executive Intelligence Summary (Sheet 7).

This sheet synthesises available quantitative and metadata signals from the
export run into a structured qualitative briefing.  Because the XBRL-based
extraction pipeline does not ingest free-text MD&A or strategic-note sections,
this sheet surfaces the following signals extracted programmatically from the
financial line items:

Sections rendered
──────────────────
  1. Export Run Metadata      — company, ticker, filing date, job ID,
                                export timestamp, reporting standard
  2. Top Revenue Drivers      — 5 largest IS revenue-band items by USD value
                                (most recent period with data)
  3. Major Cost Components    — 5 largest cost/opex items by USD value
  4. Balance Sheet Highlights — key BS metrics (assets, liabilities, equity)
  5. Period Coverage Summary  — list of all covered (year, period) pairs with
                                IS / BS / CF item counts

Layout
──────
  Two-column label–value tables within each section, separated by slate
  header rows.  Consistent navy/slate/zebra palette.

Milestone: B6 — Advanced Excel Sheet Completion.
"""

from __future__ import annotations

from decimal import Decimal
from typing import Final

_REVENUE_TAGS: Final[frozenset[str]] = frozenset({
    "us-gaap:Revenues",
    "us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax",
    "ifrs-full:Revenue",
    "ifrs-full:RevenueFromContractsWithCustomers",
    "ind-as:Revenue",
    "ind-as:RevenueFromOperations",
    "ind-as:TotalIncome",
    "ind-as:OtherIncome",
})

_COST_TAGS: Final[frozenset[str]] = frozenset({
    "us-gaap:CostOfRevenue",
    "us-gaap:CostOfGoodsSold",
    "us-gaap:OperatingExpenses",
    "us-gaap:ResearchAndDevelopmentExpense",
    "us-gaap:SellingGeneralAndAdministrativeExpense",
    "us-gaap:GeneralAndAdministrativeExpense",
    "ifrs-full:CostOfSales",
    "ifrs-full:AdministrativeExpense",
    "ifrs-full:DistributionCosts",
    "ind-as:CostOfMaterialsConsumed",
    "ind-as:PurchasesOfStockInTrade",
    "ind-as:EmployeeBenefitsExpense",
    "ind-as:FinanceCosts",
    "ind-as:OtherExpenses",
    "ind-as:Expenses",
})

_BS_HIGHLIGHT_TAGS: Final[list[tuple[str, list[str]]]] = [
    ("Total Assets",       ["us-gaap:Assets",            "ifrs-full:Assets",            "ind-as:Assets"]),
    ("Current Assets",     ["us-gaap:AssetsCurrent",     "ifrs-full:CurrentAssets",     "ind-as:CurrentAssets"]),
    ("Total Liabilities",  ["us-gaap:Liabilities",       "ifrs-full:Liabilities",       "ind-as:Liabilities"]),
    ("Current Liabilities",["us-gaap:LiabilitiesCurrent","ifrs-full:CurrentLiabilities","ind-as:CurrentLiabilities"]),
    ("Total Equity",       ["us-gaap:StockholdersEquity","ifrs-full:Equity",            "ind-as:Equity"]),
    ("Cash & Equivalents", [
        "us-gaap:CashAndCashEquivalentsAtCarryingValue",
        "ifrs-full:CashAndCashEquivalents",
        "ind-as:CashAndCashEquivalents",
    ]),
]


def _fmt_usd(value: Decimal | None) -> str:
    """Format a USD Decimal as a compact millions string, e.g. '$1,234.5 M'."""
    if value is None:
        return "—"
    try:
        millions = float(value) / 1_000_000
        return f"${millions:,.1f} M"
    except Exception:
        return str(value)


def write_strategic_initiatives_sheet(ws: object, ctx: object) -> None:  # type: ignore[type-arg]
    """
    Populate the Strategic Initiatives / Executive Summary worksheet from *ctx*.

    Called synchronously from ``ExcelExportService._build_workbook``.
    """
    from openpyxl.styles import Alignment  # type: ignore[import]
    from openpyxl.utils import get_column_letter  # type: ignore[import]

    from services.export.excel_generator import (  # type: ignore[import]
        _CLR_NAVY,
        _CLR_NAVY_LIGHT,
        _CLR_SLATE,
        _CLR_ZEBRA_EVEN,
        _CLR_WHITE,
        _CLR_BLACK,
        _make_fill,
        _make_font,
        _apply_row_fill,
        _apply_row_font,
        _period_label,
        ExportContext,
    )

    ctx_typed: ExportContext = ctx  # type: ignore[assignment]
    periods = ctx_typed.periods

    TOTAL_COLS = 4  # Label | Value | (extra) | (extra)

    fill_navy       = _make_fill(_CLR_NAVY)
    fill_navy_light = _make_fill(_CLR_NAVY_LIGHT)
    fill_slate      = _make_fill(_CLR_SLATE)
    fill_even       = _make_fill(_CLR_ZEBRA_EVEN)
    fill_white      = _make_fill(_CLR_WHITE)

    font_banner = _make_font(size=13, bold=True,  color="FFFFFF")
    font_sec    = _make_font(size=11, bold=True,  color=_CLR_BLACK)
    font_label  = _make_font(size=10, bold=True,  color=_CLR_BLACK)
    font_value  = _make_font(size=10, bold=False, color=_CLR_BLACK)
    font_note   = _make_font(size=9,  bold=False, color="595959")

    left_align   = Alignment(horizontal="left",  vertical="center", wrap_text=True)
    right_align  = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center",vertical="center")

    row = 1  # current write cursor

    def _section_header(title: str) -> None:
        nonlocal row
        _apply_row_fill(ws, row, TOTAL_COLS, fill_slate)
        c = ws.cell(row=row, column=1, value=title)  # type: ignore[union-attr]
        c.font = font_sec; c.fill = fill_slate; c.alignment = left_align
        row += 1

    def _kv_row(label: str, value: str, even: bool) -> None:
        nonlocal row
        rf = fill_even if even else fill_white
        _apply_row_fill(ws, row, TOTAL_COLS, rf)
        lc = ws.cell(row=row, column=1, value=label)  # type: ignore[union-attr]
        lc.font = font_label; lc.fill = rf; lc.alignment = left_align
        vc = ws.cell(row=row, column=2, value=value)  # type: ignore[union-attr]
        vc.font = font_value; vc.fill = rf; vc.alignment = left_align
        row += 1

    def _blank_row() -> None:
        nonlocal row
        _apply_row_fill(ws, row, TOTAL_COLS, fill_white)
        row += 1

    # ── Title banner ──────────────────────────────────────────────────────────
    _apply_row_fill(ws, row, TOTAL_COLS, fill_navy)
    tc = ws.cell(row=row, column=1,  # type: ignore[union-attr]
        value=(
            f"Executive Intelligence Summary — {ctx_typed.company_name}"
            f" ({ctx_typed.company_ticker})  ·  {ctx_typed.period_range_label}"
        ),
    )
    tc.font = font_banner; tc.fill = fill_navy; tc.alignment = left_align
    row += 1

    # ── Note ──────────────────────────────────────────────────────────────────
    _apply_row_fill(ws, row, TOTAL_COLS, fill_slate)
    nc = ws.cell(row=row, column=1,  # type: ignore[union-attr]
        value=(
            "Auto-generated from structured XBRL extraction data.  "
            "All monetary values in USD.  Qualitative MD&A commentary requires "
            "manual review of source filings."
        ),
    )
    nc.font = font_note; nc.fill = fill_slate; nc.alignment = left_align
    row += 1
    _blank_row()

    # ── Section 1: Export Run Metadata ────────────────────────────────────────
    _section_header("1. Export Run Metadata")
    meta_rows = [
        ("Company",             ctx_typed.company_name),
        ("Ticker",              ctx_typed.company_ticker),
        ("Reporting Standard",  ctx_typed.reporting_standard),
        ("Primary Fiscal Year", str(ctx_typed.fiscal_year)),
        ("Primary Period",      ctx_typed.fiscal_period),
        ("Period Range",        ctx_typed.period_range_label),
        ("Filing Date",         str(ctx_typed.filing_date) if ctx_typed.filing_date else "—"),
        ("Export Timestamp",    ctx_typed.export_timestamp.strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("Job ID",              str(ctx_typed.job_id)),
        ("Total Line Items",    str(
            len(ctx_typed.is_items) + len(ctx_typed.bs_items) + len(ctx_typed.cf_items)
        )),
    ]
    for i, (lbl, val) in enumerate(meta_rows):
        _kv_row(lbl, val, i % 2 == 0)
    _blank_row()

    # ── Section 2: Top Revenue Drivers ────────────────────────────────────────
    _section_header("2. Top Revenue Drivers")
    # Find most recent period, get top-5 revenue-band items by USD value
    rev_items = [
        i for i in ctx_typed.is_items
        if i.canonical_field in _REVENUE_TAGS and i.value_usd is not None
    ]
    # Use most recent period
    if periods:
        latest = periods[-1]
        rev_items_latest = [
            i for i in rev_items if i.period_key == latest
        ]
    else:
        rev_items_latest = rev_items

    rev_items_sorted = sorted(
        rev_items_latest, key=lambda x: abs(x.value_usd or Decimal(0)), reverse=True
    )[:5]

    if rev_items_sorted:
        for i, item in enumerate(rev_items_sorted):
            period_label = _period_label(*item.period_key)
            _kv_row(
                f"{item.concept_label}",
                f"{_fmt_usd(item.value_usd)}  [{period_label}]",
                i % 2 == 0,
            )
    else:
        _kv_row("No revenue items found", "—", True)
    _blank_row()

    # ── Section 3: Major Cost Components ─────────────────────────────────────
    _section_header("3. Major Cost Components")
    cost_items = [
        i for i in ctx_typed.is_items
        if i.canonical_field in _COST_TAGS and i.value_usd is not None
    ]
    if periods:
        cost_latest = [i for i in cost_items if i.period_key == periods[-1]]
    else:
        cost_latest = cost_items

    cost_sorted = sorted(
        cost_latest, key=lambda x: abs(x.value_usd or Decimal(0)), reverse=True
    )[:5]

    if cost_sorted:
        for i, item in enumerate(cost_sorted):
            period_label = _period_label(*item.period_key)
            _kv_row(
                f"{item.concept_label}",
                f"{_fmt_usd(item.value_usd)}  [{period_label}]",
                i % 2 == 0,
            )
    else:
        _kv_row("No cost items found", "—", True)
    _blank_row()

    # ── Section 4: Balance Sheet Highlights ───────────────────────────────────
    _section_header("4. Balance Sheet Highlights")
    # Build BS pivot for most recent period
    if periods:
        latest_bs = periods[-1]
        bs_period_data: dict[str, Decimal] = {}
        for item in ctx_typed.bs_items:
            if item.period_key == latest_bs and item.value_usd is not None:
                bs_period_data[item.canonical_field] = item.value_usd

        period_lbl = _period_label(*latest_bs)
        for i, (concept_name, tag_list) in enumerate(_BS_HIGHLIGHT_TAGS):
            value: Decimal | None = None
            for tag in tag_list:
                v = bs_period_data.get(tag)
                if v is not None:
                    value = v
                    break
            _kv_row(
                f"{concept_name}  [{period_lbl}]",
                _fmt_usd(value),
                i % 2 == 0,
            )
    else:
        _kv_row("No balance sheet data", "—", True)
    _blank_row()

    # ── Section 5: Period Coverage Summary ────────────────────────────────────
    _section_header("5. Period Coverage Summary")
    if periods:
        is_by_period  = {}
        bs_by_period  = {}
        cf_by_period  = {}
        for item in ctx_typed.is_items:
            is_by_period[item.period_key] = is_by_period.get(item.period_key, 0) + 1
        for item in ctx_typed.bs_items:
            bs_by_period[item.period_key] = bs_by_period.get(item.period_key, 0) + 1
        for item in ctx_typed.cf_items:
            cf_by_period[item.period_key] = cf_by_period.get(item.period_key, 0) + 1

        # Header sub-row
        _apply_row_fill(ws, row, TOTAL_COLS, fill_navy_light)
        for col, hdr in [(1, "Period"), (2, "IS Rows"), (3, "BS Rows"), (4, "CF Rows")]:
            c = ws.cell(row=row, column=col, value=hdr)  # type: ignore[union-attr]
            c.font = font_label; c.fill = fill_navy_light; c.alignment = center_align
        row += 1

        for i, period in enumerate(periods):
            rf = fill_even if (i % 2 == 0) else fill_white
            _apply_row_fill(ws, row, TOTAL_COLS, rf)
            plbl = _period_label(*period)
            for col, val in [
                (1, plbl),
                (2, str(is_by_period.get(period, 0))),
                (3, str(bs_by_period.get(period, 0))),
                (4, str(cf_by_period.get(period, 0))),
            ]:
                c = ws.cell(row=row, column=col, value=val)  # type: ignore[union-attr]
                c.font = font_value; c.fill = rf
                c.alignment = right_align if col > 1 else left_align
            row += 1
    else:
        _kv_row("No periods found", "—", True)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = 42  # type: ignore[union-attr]
    ws.column_dimensions[get_column_letter(2)].width = 32  # type: ignore[union-attr]
    ws.column_dimensions[get_column_letter(3)].width = 14  # type: ignore[union-attr]
    ws.column_dimensions[get_column_letter(4)].width = 14  # type: ignore[union-attr]
