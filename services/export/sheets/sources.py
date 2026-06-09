"""
Sheet 8 — Mandatory Regulatory Audit Log.

Amendment V1.2, Section 6.1 — Mandatory Regulatory Audit Log:
  Every Excel workbook MUST contain an audit log sheet with one row per
  extracted data point, providing:

    Column A — Source URL       : Hyperlink to the SEC filing document (clickable).
    Column B — Concept Tag      : Canonical XBRL field name or normalised label.
    Column C — SHA-256 Hash     : Hex digest of the source document (64 chars).
    Column D — Extraction Method: How the value was extracted (xbrl/pdf/ocr/ai).
    Column E — Extraction Timestamp: ISO 8601 UTC timestamp.
    Column F — Page Number      : PDF page where the value was found (AI/PDF only).
    Column G — Bounding Box     : "x_min,y_min,x_max,y_max" string (AI/PDF only).
    Column H — Fiscal Year      : The fiscal year of the data point.
    Column I — Fiscal Period    : Q1/Q2/Q3/Q4/FY.
    Column J — Reporting Standard: US_GAAP | IFRS | IND_AS.

  This sheet:
    - Is ALWAYS the 8th sheet in the workbook (index position 7).
    - Is PROTECTED (read-only for end users) to preserve audit integrity.
    - MUST NOT be deleted or hidden.

Amendment V1.2, Section 6.2 — Data Quality Watermarking:
  AI-extracted cells elsewhere in the workbook must carry a cell comment:
      [Lineage Context: AI-Assisted Non-XBRL Extraction |
       Document Source: Page {page_number} |
       Confidence: {confidence_pct}%]

  This sheet provides the backing data; the ``watermark_ai_cells`` function
  below applies the comments to individual data cells in other sheets.

Milestone: M6 (skeleton pre-provisioned by Amendment V1.2 compliance sweep)
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime


# ---------------------------------------------------------------------------
# Audit row data class
# ---------------------------------------------------------------------------


@dataclass
class AuditRow:
    """
    One row in the Sheet 8 audit log — represents one extracted data point.

    Attributes:
        source_url:           URL to the SEC filing document. Will be rendered
                              as a clickable hyperlink in the Excel sheet.
        concept_tag:          Canonical field name (e.g. 'us-gaap:Revenues').
        sha256_hash:          SHA-256 hex digest of the source document.
        extraction_method:    'xbrl' | 'pdf' | 'ocr' | 'ai'.
        extraction_timestamp: ISO 8601 UTC timestamp of extraction.
        fiscal_year:          Fiscal year integer (e.g. 2023).
        fiscal_period:        Q1 | Q2 | Q3 | Q4 | FY.
        reporting_standard:   US_GAAP | IFRS | IND_AS.
        page_number:          PDF page number. None for XBRL-extracted values.
        bounding_box:         "x_min,y_min,x_max,y_max". None for XBRL.
        confidence_pct:       AI model confidence score. None for XBRL/PDF.
    """

    source_url: str
    concept_tag: str
    sha256_hash: str
    extraction_method: str
    extraction_timestamp: datetime
    fiscal_year: int
    fiscal_period: str
    reporting_standard: str
    page_number: int | None = None
    bounding_box: str | None = None  # "x_min,y_min,x_max,y_max"
    confidence_pct: float | None = None


# ---------------------------------------------------------------------------
# Sheet 8 column specification
# ---------------------------------------------------------------------------

_COLUMNS = [
    ("A", "Source URL",            28),
    ("B", "Concept Tag",           36),
    ("C", "SHA-256 Hash",          68),
    ("D", "Extraction Method",     18),
    ("E", "Extraction Timestamp",  22),
    ("F", "Page Number",           12),
    ("G", "Bounding Box",          26),
    ("H", "Fiscal Year",           12),
    ("I", "Fiscal Period",         14),
    ("J", "Reporting Standard",    18),
]

# Header background colour — dark regulatory blue for audit sheets.
_HEADER_FILL_HEX = "1F3864"
_HEADER_FONT_HEX = "FFFFFF"
_SHEET_TAB_COLOR = "C00000"   # Red tab signals mandatory regulatory content.


def build_audit_log_sheet(
    workbook: object,
    rows: list[AuditRow],
    *,
    sheet_name: str = "Audit Log (Sheet 8)",
    protect: bool = True,
) -> object:
    """
    Build and return the Sheet 8 Mandatory Regulatory Audit Log.

    Creates a new worksheet in ``workbook`` at position 8 (index 7),
    writes the audit rows with source URL hyperlinks, and optionally
    protects the sheet to prevent accidental modification.

    Args:
        workbook:   openpyxl Workbook instance.
        rows:       List of AuditRow objects — one per extracted data point.
        sheet_name: Display name of the sheet (default: "Audit Log (Sheet 8)").
        protect:    When True, the sheet is protected against editing by end
                    users. The audit log must not be modifiable (Amendment §6.1).

    Returns:
        The created openpyxl Worksheet object.

    Note:
        This function requires openpyxl. At M4/M5 the function is imported
        and called by the export builder (services/export/builder.py) which
        manages the openpyxl dependency. The module-level import is deferred
        to avoid a hard dependency at import time for code paths that do not
        exercise the export layer.
    """
    import openpyxl  # type: ignore[import]  # deferred — not required at M4
    from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import]
    from openpyxl.utils import get_column_letter  # type: ignore[import]

    ws = workbook.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = _SHEET_TAB_COLOR

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL_HEX)
    header_font = Font(color=_HEADER_FONT_HEX, bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col_idx, (col_letter, header_text, col_width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[col_letter].width = col_width

    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, audit_row in enumerate(rows, start=2):
        # Column A — Source URL as hyperlink.
        url_cell = ws.cell(row=row_idx, column=1, value=audit_row.source_url)
        url_cell.hyperlink = audit_row.source_url
        url_cell.style = "Hyperlink"

        ws.cell(row=row_idx, column=2, value=audit_row.concept_tag)
        ws.cell(row=row_idx, column=3, value=audit_row.sha256_hash)
        ws.cell(row=row_idx, column=4, value=audit_row.extraction_method)
        ws.cell(
            row=row_idx,
            column=5,
            value=audit_row.extraction_timestamp.isoformat(timespec="seconds") + "Z",
        )
        ws.cell(row=row_idx, column=6, value=audit_row.page_number)
        ws.cell(row=row_idx, column=7, value=audit_row.bounding_box)
        ws.cell(row=row_idx, column=8, value=audit_row.fiscal_year)
        ws.cell(row=row_idx, column=9, value=audit_row.fiscal_period)
        ws.cell(row=row_idx, column=10, value=audit_row.reporting_standard)

    # ── Sheet protection (Amendment V1.2 §6.1) ───────────────────────────────
    if protect:
        ws.protection.sheet = True
        ws.protection.enable()

    return ws


# ---------------------------------------------------------------------------
# Amendment V1.2 §6.2 — AI cell watermarking
# ---------------------------------------------------------------------------


def watermark_ai_cells(
    worksheet: object,
    cell_coordinate: str,
    *,
    page_number: int,
    confidence_pct: float,
    author: str = "Financial Data Hub",
) -> None:
    """
    Add a lineage context comment to an AI-extracted cell in any worksheet.

    Amendment V1.2 §6.2 — Data Quality Watermarking:
    AI-extracted cell comments use the format:
        [Lineage Context: AI-Assisted Non-XBRL Extraction |
         Document Source: Page {page_number} |
         Confidence: {confidence_pct:.0f}%]

    Args:
        worksheet:        openpyxl Worksheet containing the target cell.
        cell_coordinate:  Cell address string, e.g. 'B5'.
        page_number:      PDF page number where the value was extracted.
        confidence_pct:   AI model confidence score (0–100).
        author:           Comment author displayed in Excel.
    """
    from openpyxl.comments import Comment  # type: ignore[import]

    lineage_text = (
        f"[Lineage Context: AI-Assisted Non-XBRL Extraction | "
        f"Document Source: Page {page_number} | "
        f"Confidence: {confidence_pct:.0f}%]"
    )
    comment = Comment(text=lineage_text, author=author)
    worksheet[cell_coordinate].comment = comment
