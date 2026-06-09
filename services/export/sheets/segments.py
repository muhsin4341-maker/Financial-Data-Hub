"""
Sheet builder: Segments — Revenue & Cost Contribution Analysis (Sheet 6).

Because the platform's extraction pipeline ingests standardised XBRL line
items rather than segment-note footnote tables, this sheet performs a
structural revenue-contribution analysis using the canonical Income Statement
items already present in ``ExportContext``.  It groups line items into four
logical bands and surfaces each line's share-of-revenue percentage across all
chronological fiscal periods.

Bands rendered (in order)
──────────────────────────
  1. Revenue Sources        — all canonical revenue tags
  2. Cost of Sales          — cost-of-revenue / cost-of-goods-sold tags
  3. Operating Expenses     — R&D, SG&A, D&A, other opex tags
  4. Income Milestones      — Gross Profit, Operating Income, Net Income, EBT

Layout
──────
  Row 1:   Title banner (navy)
  Row 2:   Methodology note (slate)
  Row 3:   Blank separator
  Row 4:   Column headers — Concept | Canonical Tag | Band |
             then per period: "FY 2022 Value (USD)" | "FY 2022 % Revenue"
  Row 5+:  Band separator rows (slate) + data rows (zebra)

Number formats
──────────────
  USD values : "#,##0"   (whole-dollar, no decimals)
  % Revenue  : "0.0%"    (stored as 0–1 fraction)

Missing values render as "—".

Milestone: B6 — Advanced Excel Sheet Completion.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from typing import Final

# ---------------------------------------------------------------------------
# Band membership
# ---------------------------------------------------------------------------

_BAND_REVENUE: Final[str] = "Revenue Sources"
_BAND_COGS:    Final[str] = "Cost of Sales"
_BAND_OPEX:    Final[str] = "Operating Expenses"
_BAND_INCOME:  Final[str] = "Income Milestones"

_TAG_TO_BAND: Final[dict[str, str]] = {
    # Revenue
    "us-gaap:Revenues":                                              _BAND_REVENUE,
    "us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax":   _BAND_REVENUE,
    "ifrs-full:Revenue":                                             _BAND_REVENUE,
    "ifrs-full:RevenueFromContractsWithCustomers":                   _BAND_REVENUE,
    "ind-as:Revenue":                                                _BAND_REVENUE,
    "ind-as:RevenueFromOperations":                                  _BAND_REVENUE,
    "ind-as:TotalIncome":                                            _BAND_REVENUE,
    "ind-as:OtherIncome":                                            _BAND_REVENUE,
    # Cost of Sales
    "us-gaap:CostOfRevenue":                                         _BAND_COGS,
    "us-gaap:CostOfGoodsSold":                                       _BAND_COGS,
    "ifrs-full:CostOfSales":                                         _BAND_COGS,
    "ind-as:CostOfMaterialsConsumed":                                _BAND_COGS,
    "ind-as:PurchasesOfStockInTrade":                                _BAND_COGS,
    "ind-as:ChangesInInventoriesOfFinishedGoodsWorkInProgressAndStockInTrade": _BAND_COGS,
    # Operating Expenses
    "us-gaap:ResearchAndDevelopmentExpense":                         _BAND_OPEX,
    "us-gaap:SellingGeneralAndAdministrativeExpense":                _BAND_OPEX,
    "us-gaap:GeneralAndAdministrativeExpense":                       _BAND_OPEX,
    "us-gaap:SellingExpense":                                        _BAND_OPEX,
    "us-gaap:MarketingExpense":                                      _BAND_OPEX,
    "us-gaap:DepreciationDepletionAndAmortization":                  _BAND_OPEX,
    "us-gaap:OperatingExpenses":                                     _BAND_OPEX,
    "ifrs-full:DistributionCosts":                                   _BAND_OPEX,
    "ifrs-full:AdministrativeExpense":                               _BAND_OPEX,
    "ifrs-full:ResearchAndDevelopmentExpense":                       _BAND_OPEX,
    "ind-as:EmployeeBenefitsExpense":                                _BAND_OPEX,
    "ind-as:FinanceCosts":                                           _BAND_OPEX,
    "ind-as:DepreciationDepletionAndAmortisation":                   _BAND_OPEX,
    "ind-as:OtherExpenses":                                          _BAND_OPEX,
    "ind-as:Expenses":                                               _BAND_OPEX,
    # Income milestones
    "us-gaap:GrossProfit":                                           _BAND_INCOME,
    "ifrs-full:GrossProfit":                                         _BAND_INCOME,
    "ind-as:GrossProfit":                                            _BAND_INCOME,
    "us-gaap:OperatingIncomeLoss":                                   _BAND_INCOME,
    "ifrs-full:ProfitLossFromOperatingActivities":                   _BAND_INCOME,
    "ind-as:ProfitBeforeExceptionalItemsAndTax":                     _BAND_INCOME,
    "us-gaap:IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest": _BAND_INCOME,
    "ifrs-full:ProfitLossBeforeTax":                                 _BAND_INCOME,
    "ind-as:ProfitBeforeTax":                                        _BAND_INCOME,
    "us-gaap:NetIncomeLoss":                                         _BAND_INCOME,
    "ifrs-full:ProfitLoss":                                          _BAND_INCOME,
    "ind-as:ProfitLoss":                                             _BAND_INCOME,
}

_BAND_ORDER: Final[list[str]] = [
    _BAND_REVENUE,
    _BAND_COGS,
    _BAND_OPEX,
    _BAND_INCOME,
]

_REVENUE_TAGS: Final[frozenset[str]] = frozenset({
    "us-gaap:Revenues",
    "us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax",
    "ifrs-full:Revenue",
    "ifrs-full:RevenueFromContractsWithCustomers",
    "ind-as:Revenue",
    "ind-as:RevenueFromOperations",
    "ind-as:TotalIncome",
})


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def write_segments_sheet(ws: object, ctx: object) -> None:  # type: ignore[type-arg]
    """
    Populate the Segments / Revenue Contribution worksheet from *ctx*.

    Called synchronously from ``ExcelExportService._build_workbook``.
    """
    from openpyxl.styles import Alignment  # type: ignore[import]
    from openpyxl.utils import get_column_letter  # type: ignore[import]

    from services.export.excel_generator import (  # type: ignore[import]
        _CLR_NAVY,
        _CLR_SLATE,
        _CLR_ZEBRA_EVEN,
        _CLR_WHITE,
        _CLR_BLACK,
        _make_fill,
        _make_font,
        _apply_row_fill,
        _apply_row_font,
        _auto_fit_columns,
        _period_label,
        ExportContext,
    )

    ctx_typed: ExportContext = ctx  # type: ignore[assignment]
    periods = ctx_typed.periods

    # ── Build USD pivot ───────────────────────────────────────────────────────
    pivot: dict[tuple[int, str], dict[str, Decimal]] = {}
    for item in ctx_typed.is_items:
        if item.value_usd is None:
            continue
        pivot.setdefault(item.period_key, {})[item.canonical_field] = item.value_usd

    def _get_revenue(period: tuple[int, str]) -> Decimal | None:
        pd_data = pivot.get(period, {})
        for tag in _REVENUE_TAGS:
            v = pd_data.get(tag)
            if v is not None:
                return v
        return None

    # ── Deduplicate IS items to one row per canonical_field ───────────────────
    seen: dict[str, object] = {}
    for item in ctx_typed.is_items:
        cf = item.canonical_field
        if cf not in _TAG_TO_BAND:
            continue
        existing = seen.get(cf)
        if existing is None or item.fiscal_year > existing.fiscal_year:  # type: ignore[attr-defined]
            seen[cf] = item

    bands: dict[str, list[object]] = {b: [] for b in _BAND_ORDER}
    for item in ctx_typed.is_items:
        cf = item.canonical_field
        band = _TAG_TO_BAND.get(cf)
        if band and item is seen.get(cf):
            bands[band].append(item)

    # ── Column layout ─────────────────────────────────────────────────────────
    _COL_LABEL = 1
    _COL_TAG   = 2
    _COL_BAND  = 3
    _COL_DATA  = 4
    total_cols = (_COL_DATA + len(periods) * 2 - 1) if periods else (_COL_DATA + 1)

    fill_navy  = _make_fill(_CLR_NAVY)
    fill_slate = _make_fill(_CLR_SLATE)
    fill_even  = _make_fill(_CLR_ZEBRA_EVEN)
    fill_white = _make_fill(_CLR_WHITE)

    font_header    = _make_font(size=11, bold=True,  color="FFFFFF")
    font_sub       = _make_font(size=10, bold=False, color=_CLR_BLACK)
    font_label_b   = _make_font(size=10, bold=True,  color=_CLR_BLACK)
    font_separator = _make_font(size=10, bold=True,  color=_CLR_BLACK)
    font_data      = _make_font(size=10, bold=False, color=_CLR_BLACK)
    font_note      = _make_font(size=9,  bold=False, color="595959")

    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")
    right_align  = Alignment(horizontal="right",  vertical="center")

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    _apply_row_fill(ws, 1, total_cols, fill_navy)
    _apply_row_font(ws, 1, total_cols, font_header)
    tc = ws.cell(row=1, column=_COL_LABEL,  # type: ignore[union-attr]
        value=(
            f"Revenue & Cost Contribution — {ctx_typed.company_name}"
            f" ({ctx_typed.company_ticker})  ·  {ctx_typed.period_range_label}"
        ),
    )
    tc.font = font_header; tc.fill = fill_navy; tc.alignment = left_align

    # ── Row 2: Note ───────────────────────────────────────────────────────────
    _apply_row_fill(ws, 2, total_cols, fill_slate)
    nc = ws.cell(row=2, column=_COL_LABEL,  # type: ignore[union-attr]
        value=(
            "USD-translated values.  % Revenue = line value ÷ period total revenue. "
            "Geographic / product-level footnote tables require direct XBRL segment tags."
        ),
    )
    nc.font = font_note; nc.fill = fill_slate; nc.alignment = left_align

    # ── Row 3: Blank ──────────────────────────────────────────────────────────
    _apply_row_fill(ws, 3, total_cols, fill_white)

    # ── Row 4: Column headers ─────────────────────────────────────────────────
    _apply_row_fill(ws, 4, total_cols, fill_navy)
    _apply_row_font(ws, 4, total_cols, font_header)
    for col, lbl in [(_COL_LABEL, "Concept"), (_COL_TAG, "Canonical Tag"), (_COL_BAND, "Band")]:
        c = ws.cell(row=4, column=col, value=lbl)  # type: ignore[union-attr]
        c.font = font_header; c.fill = fill_navy; c.alignment = center_align

    for p_idx, period in enumerate(periods):
        base = _COL_DATA + p_idx * 2
        plbl = _period_label(*period)
        for off, sub in [(0, "Value (USD)"), (1, "% Revenue")]:
            c = ws.cell(row=4, column=base + off, value=f"{plbl} {sub}")  # type: ignore[union-attr]
            c.font = font_header; c.fill = fill_navy; c.alignment = center_align

    # ── Data rows ─────────────────────────────────────────────────────────────
    row_idx = 5
    any_data = False

    for band_name in _BAND_ORDER:
        items_in_band = bands[band_name]
        if not items_in_band:
            continue
        any_data = True

        # Separator
        _apply_row_fill(ws, row_idx, total_cols, fill_slate)
        sep = ws.cell(row=row_idx, column=_COL_LABEL, value=band_name)  # type: ignore[union-attr]
        sep.font = font_separator; sep.fill = fill_slate; sep.alignment = left_align
        row_idx += 1
        data_row = 0

        for item in items_in_band:  # type: ignore[attr-defined]
            row_fill = fill_even if (data_row % 2 == 0) else fill_white
            _apply_row_fill(ws, row_idx, total_cols, row_fill)

            lc = ws.cell(row=row_idx, column=_COL_LABEL, value=item.concept_label)  # type: ignore[union-attr]
            lc.font = font_label_b; lc.fill = row_fill; lc.alignment = left_align

            tg = ws.cell(row=row_idx, column=_COL_TAG, value=item.canonical_field)  # type: ignore[union-attr]
            tg.font = font_sub; tg.fill = row_fill; tg.alignment = left_align

            bd = ws.cell(row=row_idx, column=_COL_BAND, value=band_name)  # type: ignore[union-attr]
            bd.font = font_sub; bd.fill = row_fill; bd.alignment = left_align

            for p_idx, period in enumerate(periods):
                base    = _COL_DATA + p_idx * 2
                pd_data = pivot.get(period, {})
                value   = pd_data.get(item.canonical_field)  # type: ignore[attr-defined]
                revenue = _get_revenue(period)

                # Value cell
                if value is None:
                    vc = ws.cell(row=row_idx, column=base, value="—")  # type: ignore[union-attr]
                    vc.font = font_data; vc.fill = row_fill; vc.alignment = center_align
                else:
                    vc = ws.cell(row=row_idx, column=base)  # type: ignore[union-attr]
                    vc.value = float(value); vc.number_format = "#,##0"
                    vc.font = font_data; vc.fill = row_fill; vc.alignment = right_align

                # % Revenue cell
                pc = ws.cell(row=row_idx, column=base + 1)  # type: ignore[union-attr]
                if value is None or revenue is None or revenue == Decimal(0):
                    pc.value = "—"; pc.alignment = center_align
                else:
                    try:
                        pct = float(value / revenue)
                        pc.value = pct; pc.number_format = "0.0%"; pc.alignment = right_align
                    except (InvalidOperation, ZeroDivisionError):
                        pc.value = "—"; pc.alignment = center_align
                pc.font = font_data; pc.fill = row_fill

            row_idx  += 1
            data_row += 1

    if not any_data:
        ec = ws.cell(row=row_idx, column=_COL_LABEL,  # type: ignore[union-attr]
                     value="No Income Statement data available.")
        ec.font = _make_font(size=10, color="808080"); ec.alignment = left_align

    # ── Column widths ─────────────────────────────────────────────────────────
    _auto_fit_columns(ws)
    ws.column_dimensions[get_column_letter(_COL_LABEL)].width = 38  # type: ignore[union-attr]
    ws.column_dimensions[get_column_letter(_COL_TAG)].width   = 54  # type: ignore[union-attr]
    ws.column_dimensions[get_column_letter(_COL_BAND)].width  = 22  # type: ignore[union-attr]
